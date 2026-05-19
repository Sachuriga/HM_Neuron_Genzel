"""
Rat Hex Maze Preprocessing — Step 1 + Step 2 with Homeboxes

Finds every .xlsx in EXCEL_FOLDER, processes all rows, and saves one
results Excel file per input file into OUTPUT_FOLDER.

The output is a copy of the original file with the computed columns
appended to the right of the data sheet. All other worksheets and all
existing cell colours are preserved. Only the newly added cells in
flagged rows are coloured red.

Computed columns added:
  Step 1 — shortest_path, eat_on_1_encounter, n_nodes_visited, food_reached,
            dist_tra, dt_rel_sp, dt_min_sp, dir_run_mat_perf,
            node_choices_binary, perc_corr_choices
  Step 2 — node_island_in, island_short_path, island_dt_traveled, perf_in_island
  Errors — flag  (non-empty = row skipped, new cells highlighted red)

Requirements: pip install networkx pandas openpyxl
"""

import argparse
import glob
import os
import shutil
import numpy as np
import pandas as pd
import networkx as nx
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ── Parse folder from command line ───────────────────────────────────────────
parser = argparse.ArgumentParser(description='Rat Hex Maze Analysis')
parser.add_argument('--input_folder',  '-i', required=True,
                    help='Input folder (ip[n]) containing .xlsx files')
parser.add_argument('--output_folder', '-o', required=True,
                    help='Output folder (op[n]) where results are saved')
args = parser.parse_args()

FOLDER_PAIRS = [(args.input_folder, args.output_folder)]
print(f'Input:  {args.input_folder}')
print(f'Output: {args.output_folder}')

# ── Graph: 96 maze nodes + 2 homeboxes (501, 502) ────────────────────────────
NDS = (list(range(101, 125)) + list(range(201, 225)) +
       list(range(301, 325)) + list(range(401, 425)) + [501, 502])

_s1 = [1,1,2,3,3,4,5,6,6,7,8,8,9,10,10,11,12,13,14,14,15,16,16,17,18,18,20,21,22,23]
_t1 = [2,7,3,4,9,5,11,13,7,8,9,15,10,11,17,12,19,14,20,15,16,22,17,18,19,24,21,22,23,24]
_between = [(24,25),(44,53),(72,73),(21,50),(47,76),(21,97),(47,98),(50,97),(76,98)]

def _build_graph():
    G = nx.Graph()
    G.add_nodes_from(NDS)
    for offset in [0, 24, 48, 72]:
        for s, t in zip(_s1, _t1):
            G.add_edge(NDS[offset + s - 1], NDS[offset + t - 1])
    for s, t in _between:
        G.add_edge(NDS[s - 1], NDS[t - 1])
    return G

G     = _build_graph()
_DIST = dict(nx.all_pairs_shortest_path_length(G))

# Island-level graph (islands 1-4 only; homeboxes excluded)
_between = [(24,25),(44,53),(72,73),(21,50),(47,76),(21,97),(47,98),(50,97),(76,98)]
_ISLAND_G = nx.Graph()
_ISLAND_G.add_nodes_from([1, 2, 3, 4])
for _s, _t in _between:
    _n1, _n2 = NDS[_s - 1], NDS[_t - 1]
    _i1, _i2 = _n1 // 100, _n2 // 100
    if 1 <= _i1 <= 4 and 1 <= _i2 <= 4 and _i1 != _i2:
        _ISLAND_G.add_edge(_i1, _i2)
_ISLAND_DIST = dict(nx.all_pairs_shortest_path_length(_ISLAND_G))

RED_FILL = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')

# Columns this script computes and writes values for.
COMPUTED_COLS = {
    'distance_start_goal_island',
    'distance_start_goal_nodes',
    'path_length_start_goal_island_island_hit',
    'path_length_start_goal_island_node_hit',
    'path_length_start_goal_nodes_node_hit',
    'norm_path_length_start_goal_island_island_hit',
    'norm_path_length_start_goal_island_node_hit',
    'norm_path_length_start_goal_nodes_node_hit',
    'shortest_path', 'eat_on_1_encounter', 'n_nodes_visited', 'food_reached',
    'dist_tra', 'dt_rel_sp', 'dt_min_sp', 'dir_run_mat_perf',
    'node_choices_binary', 'perc_correct_choices',
    'isl_node_in', 'isl_short_path', 'isl_dt_trav', 'perf_in_island',
    'flag',
}

# All columns that appear after "comment" in the reference file, in exact order.
# Columns in COMPUTED_COLS get values written; others get an empty header created only.
ALL_OUTPUT_COLS = [
    'distance_start_goal_island',
    'distance_start_goal_nodes',
    'path_length_start_goal_island_island_hit',
    'path_length_start_goal_island_node_hit',
    'path_length_start_goal_nodes_node_hit',
    'norm_path_length_start_goal_island_island_hit',
    'norm_path_length_start_goal_island_node_hit',
    'norm_path_length_start_goal_nodes_node_hit',
    'Diff_Lat_reach_eat',
    'goal_island_i_e',
    'start_island_i_e',
    'shortest_path',
    'n_nodes_visited',
    'food_reached',
    'dist_tra',
    'dt_rel_sp',
    'dt_min_sp',
    'dir_run_mat_perf',
    'dir_run_mat_lat',
    'node_choices_binary',
    'perc_correct_choices',
    'drug',
    'number_times_drug_infused',
    'lg-DT_REL_SP',
    'lg10-DT_REL_SP',
    'isl_node_in',
    'isl_short_path',
    'isl_dt_trav',
    'perf_in_island',
    'lg_perf_I',
    'Project',
    'Training_order',
    'Implant',
    'Number_of_goal_locations',
    'flag',
]

def _pick_sheet(xl):
    for name in xl.sheet_names:
        if name.lower() == 'raw':
            return name
    return xl.sheet_names[0]

def _compute(exc_path):
    """Read the data sheet and compute all new columns. Returns (df, sheet_name)."""
    xl    = pd.ExcelFile(exc_path)
    sheet = _pick_sheet(xl)
    df    = pd.read_excel(exc_path, sheet_name=sheet, header=0)
    print(f'  Sheet: "{sheet}"  |  {len(df)} rows')

    for col in COMPUTED_COLS:
        df[col] = np.nan if col not in ('flag', 'node_choices_binary') else ''

    if 'path_to_reach' not in df.columns:
        print(f"  WARNING: column 'path_to_reach' not found — all rows flagged, skipping computation.")
        for i in df.index:
            df.at[i, 'flag'] = 'missing path_to_reach column'
        return df, sheet

    for i, row in df.iterrows():
        path_val = row.get('path_to_reach')
        if path_val is None or pd.isna(path_val) or str(path_val).strip() == '':
            comment = str(row.get('comment', '')) if pd.notna(row.get('comment')) else ''
            df.at[i, 'flag'] = comment.strip() if comment.strip() else 'unknown error'
            continue

        try:
            path       = [int(float(x)) for x in str(path_val).split(',') if x.strip()]
            start_node = int(row['start_node_n'])
            goal_node  = int(row['goal_node_n'])

            if path[0] != start_node:
                raise ValueError(f'start_node_n ({start_node}) != first path node ({path[0]})')

            unknown = [n for n in path if n not in _DIST]
            if unknown:
                raise ValueError(f'unknown node(s) in path: {unknown}')

            n_nodes      = len(path)
            food_reached = goal_node in path[-2:]
            shortest     = _DIST[start_node][goal_node]
            dist_tra     = (n_nodes - 1) if food_reached else 99

            # ── Island-level distances ───────────────────────────────────────
            start_island = int(row['start_island_n'])
            goal_island  = int(row['goal_island_n'])
            dist_isl     = _ISLAND_DIST[start_island][goal_island] + 1
            dist_nodes   = shortest + 1

            df.at[i, 'distance_start_goal_island'] = dist_isl
            df.at[i, 'distance_start_goal_nodes']  = dist_nodes

            seq_raw   = str(row['seq_islands']) if pd.notna(row['seq_islands']) else ''
            seq_items = [x.strip() for x in seq_raw.split(',') if x.strip()]
            pl_isl_node = len(seq_items)
            pl_isl_isl  = len(set(seq_items))

            df.at[i, 'path_length_start_goal_island_node_hit']    = pl_isl_node
            df.at[i, 'path_length_start_goal_island_island_hit']  = pl_isl_isl
            df.at[i, 'path_length_start_goal_nodes_node_hit']     = n_nodes
            df.at[i, 'norm_path_length_start_goal_island_island_hit'] = (
                pl_isl_isl / dist_isl if dist_isl > 0 else np.nan
            )
            df.at[i, 'norm_path_length_start_goal_island_node_hit'] = (
                pl_isl_node / dist_isl if dist_isl > 0 else np.nan
            )
            df.at[i, 'norm_path_length_start_goal_nodes_node_hit'] = (
                n_nodes / dist_nodes if dist_nodes > 0 else np.nan
            )

            # ── Step 1 ───────────────────────────────────────────────────────
            df.at[i, 'shortest_path']      = shortest
            df.at[i, 'eat_on_1_encounter'] = int(path[-1] == goal_node)
            df.at[i, 'n_nodes_visited']    = n_nodes
            df.at[i, 'food_reached']       = int(food_reached)
            df.at[i, 'dist_tra']           = dist_tra
            df.at[i, 'dt_rel_sp']          = dist_tra / shortest if shortest > 0 else np.nan
            df.at[i, 'dt_min_sp']          = dist_tra - shortest
            df.at[i, 'dir_run_mat_perf']   = int(food_reached and dist_tra == shortest)

            choices = []
            for iNode in range(n_nodes - 1):
                curr       = path[iNode]
                next_n     = path[iNode + 1]
                neighbours = list(G.neighbors(curr))
                if next_n not in neighbours:
                    choices.append(0)
                    continue
                min_sp = min(_DIST[nb][goal_node] for nb in neighbours)
                choices.append(int(_DIST[next_n][goal_node] == min_sp))

            df.at[i, 'node_choices_binary']  = ','.join(str(c) for c in choices)
            df.at[i, 'perc_correct_choices'] = (
                (sum(choices) * 100) / (n_nodes - 1) if n_nodes > 1 else np.nan
            )

            # ── Step 2 (skipped if trial is excluded) ────────────────────────
            if row['exclude_trial'] != 0:
                continue

            diffs         = [abs(path[j + 1] - path[j]) for j in range(n_nodes - 1)]
            enter_indices = [j for j, d in enumerate(diffs) if d >= 50]

            if enter_indices:
                index_enter    = enter_indices[-1]
                node_island_in = path[index_enter]
                island_sp      = _DIST[node_island_in][goal_node] + 1
                island_dt      = len(path[index_enter:])
                df.at[i, 'isl_node_in']    = node_island_in
                df.at[i, 'isl_short_path'] = island_sp
                df.at[i, 'isl_dt_trav']    = island_dt
                df.at[i, 'perf_in_island'] = island_dt / island_sp

        except Exception as e:
            df.at[i, 'flag'] = str(e)
            print(f'  Flagged row {i + 2}: {e}')

    return df, sheet

def _save(df, sheet, exc_path, out_path):
    """Copy the original file, find existing column headers by name, and write
    computed values into them. Overwrites any existing values. All other sheets,
    formatting, and cell colours are untouched. Only the newly written cells in
    flagged rows are coloured red."""
    shutil.copy2(exc_path, out_path)

    wb = load_workbook(out_path)
    ws = wb[sheet]

    # Build a map of header name → column index (1-based) from row 1
    header_to_col = {
        ws.cell(row=1, column=c).value: c
        for c in range(1, ws.max_column + 1)
    }

    # Columns that may appear under different names in different Excel files
    ALIASES = {
        'perc_correct_choices': ['perc_correct_choices', 'perc_corr_choices'],
        'isl_node_in':          ['isl_node_in',          'node_island_in'],
        'isl_short_path':       ['isl_short_path',       'island_short_path'],
        'isl_dt_trav':          ['isl_dt_trav',          'island_dt_traveled'],
    }

    # Append new columns right after the last named column, skipping phantom
    # empty columns Excel sometimes hides at the end of the sheet.
    last_named_col = max(
        (c for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value is not None),
        default=ws.max_column
    )
    next_new_col = last_named_col + 1

    # Pass 1 — walk ALL_OUTPUT_COLS in reference order.
    # Computed cols → find or create, add to col_map for writing.
    # Placeholder cols → create header only if missing, never written to.
    col_map = {}
    for col_name in ALL_OUTPUT_COLS:
        candidates = ALIASES.get(col_name, [col_name])
        matched = next((c for c in candidates if c in header_to_col), None)
        if col_name in COMPUTED_COLS:
            if matched:
                col_map[col_name] = header_to_col[matched]
            else:
                print(f'  WARNING: "{col_name}" not found in sheet — adding as new column')
                ws.cell(row=1, column=next_new_col, value=col_name)
                col_map[col_name] = next_new_col
                next_new_col += 1
        else:
            if not matched:
                ws.cell(row=1, column=next_new_col, value=col_name)
                next_new_col += 1

    # Pass 2 — pick up computed cols that live before "comment" in the input
    # section (e.g. eat_on_1_encounter) and aren't in ALL_OUTPUT_COLS.
    for col_name in COMPUTED_COLS:
        if col_name not in col_map:
            candidates = ALIASES.get(col_name, [col_name])
            matched = next((c for c in candidates if c in header_to_col), None)
            if matched:
                col_map[col_name] = header_to_col[matched]

    # Clear existing values in computed columns only
    for col_idx in col_map.values():
        for excel_row in range(2, ws.max_row + 1):
            ws.cell(row=excel_row, column=col_idx).value = None

    # Write computed values
    for i, row in df.iterrows():
        excel_row  = i + 2      # pandas index 0 → Excel row 2 (row 1 is header)
        is_flagged = bool(row['flag'])

        for col_name, col_idx in col_map.items():
            val = row[col_name]
            if isinstance(val, float) and np.isnan(val):
                val = None
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            if is_flagged:
                cell.fill = RED_FILL

    wb.save(out_path)

# ── Run over every ip/op pair ─────────────────────────────────────────────────
for ip_folder, op_folder in FOLDER_PAIRS:
    os.makedirs(op_folder, exist_ok=True)
    excel_files = [f for f in glob.glob(os.path.join(ip_folder, '*.xlsx'))
                   if not os.path.basename(f).startswith('~$')]

    if not excel_files:
        print(f'No .xlsx files found in {ip_folder}')
        continue

    print(f'\n── {ip_folder} → {op_folder}')
    for exc_path in excel_files:
        fname = os.path.basename(exc_path)
        print(f'\n  Processing: {fname}')
        df, sheet = _compute(exc_path)
        out_path  = os.path.join(op_folder, fname.replace('.xlsx', '_results.xlsx'))
        _save(df, sheet, exc_path, out_path)
        n_flagged = (df['flag'] != '').sum()
        print(f'  Saved → {out_path}  ({n_flagged} row(s) flagged red)')

print('\nAll files done.')

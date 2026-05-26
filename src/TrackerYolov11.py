# -*- coding: utf-8 -*-
'''
Title: Tracker (Headless / Excel Column Lists / Auto-Video) - UPDATED FOR YOLO11
Description: 1. Reads metadata from *RecordingMeta.xlsx (Handles vertical lists).
             2. Automatically finds 'stitched.mp4' in input_folder.
             3. Optimized for Batch/Massive Analysis.
             4. Updated to use Ultralytics YOLO11x.pt
Author: sachuriga

Based on / modified from: https://github.com/genzellab/HM_RAT
Key modifications from the original:
  - Detection backend replaced with YOLOv11 (Ultralytics)
  - Rat position uses body ('rat') class only; head class counted separately
  - Extended trial state machine: NGL variants (types 4-6), DNR logic,
    researcher-proximity triggers, force-end timers, inter-trial lockout
  - Per-trial metrics (avg speed, between-node speed, active time,
    start/end sync timestamps) written back into RecordingMeta.xlsx
  - Motion-based YOLO skip with cached bounding boxes to prevent display flash
  - Threaded video writer for non-blocking output
'''

from itertools import groupby
from datetime import datetime
from pathlib import Path
from collections import deque
from tools import mask
import cv2
from ultralytics import YOLO 
import os
import math
import time
import logging
import threading
import queue 
import numpy as np
import pandas as pd
import sys
import argparse
import glob
from tqdm import tqdm

# --- CONFIGURATION ---
FONT = cv2.FONT_HERSHEY_TRIPLEX
font = cv2.FONT_HERSHEY_PLAIN 
colors = np.random.uniform(0, 255, size=(100, 3))

def points_dist(p1, p2):
    dist = math.sqrt((p2[0] - p1[0]) ** 2 + (p2[1] - p1[1]) ** 2)
    return dist

def convert_milli(time):
    sec = (time / 1000) % 60
    minute = (time / (1000 * 60)) % 60
    hr = (time / (1000 * 60 * 60)) % 24
    return f'{int(hr):02d}:{int(minute):02d}:{sec:.3f}'

def safe_int_str(val):
    """Converts float/int to string without .0 for integers"""
    try:
        if pd.isna(val): return ""
        return str(int(float(val)))
    except:
        return str(val)

# --- CLASS: Threaded Video Writer ---
class ThreadedVideoWriter:
    def __init__(self, path, fourcc, fps, frame_size):
        self.output_file = cv2.VideoWriter(path, fourcc, fps, frame_size)
        self.queue = queue.Queue()
        self.stopped = False
        self.thread = threading.Thread(target=self.write_frames, daemon=True)
        self.thread.start()

    def write(self, frame):
        if not self.stopped:
            self.queue.put(frame)
            
    def closest_researcher_to(self, point):
        if not self.all_researchers:
            return None
        return min(self.all_researchers, key=lambda r: points_dist(r, point))
    
    def write_frames(self):
        while True:
            if self.stopped and self.queue.empty():
                break
            try:
                frame = self.queue.get(timeout=1) 
                self.output_file.write(frame)
                self.queue.task_done()
            except queue.Empty:
                continue

    def release(self):
        self.stopped = True
        self.thread.join()
        self.output_file.release()

# --- CLASS: Tracker ---
class Tracker:
    def __init__(self, vp, nl, out, metadata, onnx_weight):
        '''Tracker class initialisations'''
        self.metadata = metadata 
        self.out_path = out 
        self.model_path = onnx_weight 
        threads = list()
        
        # Load Network in main thread context to ensure model loads correctly onto GPU/CPU
        self.load_network(self.model_path)

        session = threading.Thread(target=self.load_session, args=(vp, nl, 1, out))
        threads.append(session)
        session.start()
        session.join()
            
        print('\n -Network loaded- ')

        print("Caching node dictionary...")
        self.nodes_dict = mask.create_node_dict(self.node_list)

        self._load_maze_roi(metadata.get('input_folder', out))

        self.start_nodes_locations, self.goal_locations = self.find_location(self.start_nodes, self.goal_nodes)
        print('\n  ________  SUMMARY SESSION  ________  ')
        print('\nPath video file:', self.save_video)
        print('\nTotal trials current session:', self.num_trials)
        
        self.logger = logging.getLogger('')
        self.logger.setLevel(logging.INFO)
        if self.logger.hasHandlers():
            self.logger.handlers.clear()

        logfile_name = '{}/log_{}_{}.log'.format(out, str(self.date), 'Rat' + self.rat)
        
        if not os.path.exists(out):
            os.makedirs(out, exist_ok=True)

        fh = logging.FileHandler(str(logfile_name))
        formatter = logging.Formatter('%(levelname)s : %(message)s')
        fh.setFormatter(formatter)
        self.logger.addHandler(fh)
        self.logger.info('Video Imported: {}'.format(vp))
        self.logger.info(f'The log format is: Video Timestamp(hh:mm:ss.ms), UTC Synchronised Timestamp in seconds, Rat position')
        
        print('\nCreating log files...')

        self.ts_file_loaded = False
        try:
            specific_ts_path = os.path.join(out, f'{str(self.date)}_Rat{str(self.rat)}_framewise_ts.csv')
            if os.path.exists(specific_ts_path):
                 self.sync_ts_dict = pd.read_csv(specific_ts_path, index_col=0).to_dict()
                 print("Loaded timestamp file: " + os.path.basename(specific_ts_path))
                 self.ts_file_loaded = True
            else:
                 candidates = [
                     'stitched_framewise_seconds.csv',
                     'stitched_framewise_ts.csv',
                 ]
                 loaded = False
                 for fname in candidates:
                     p = os.path.join(out, fname)
                     if os.path.exists(p):
                         print(f"Specific timestamp file not found. Loading '{fname}'...")
                         self.sync_ts_dict = pd.read_csv(p, index_col=0).to_dict()
                         self.ts_file_loaded = True
                         loaded = True
                         break
                 if not loaded:
                     raise FileNotFoundError
        except Exception:
             print("Warning: No timestamp CSV found. Logs might lack sync times.")
             self.sync_ts_dict = {"Corrected Time Stamp": {}} 

        # Inside __init__, after loading self.sync_ts_dict:
        self.ts_column_name = "Seconds From Creation" 
        if self.ts_file_loaded:
            # Check if the expected name exists, otherwise grab the first available column
            if self.ts_column_name not in self.sync_ts_dict:
                self.ts_column_name = list(self.sync_ts_dict.keys())[0]
            print(f"Using '{self.ts_column_name}' for summary timestamps.")

        self.frame_data_log = []

        self.run_vid()
    
    def change_name_csv(self, output_path):
        csvfile_name = os.path.join(output_path,f'{str(self.date)}_Rat{str(self.rat)}_framewise_ts.csv')
        stitched_name = os.path.join(output_path,'stitched_framewise_ts.csv')
        
        if os.path.exists(stitched_name):
            try:
                if os.path.exists(csvfile_name):
                    os.remove(csvfile_name)
                os.rename(stitched_name, csvfile_name)
                print(f"File renamed to: {os.path.basename(csvfile_name)}")
            except OSError as e:
                print(f"Error renaming file: {e}")
        else:
            pass
        
    def load_network(self, model_path):
        import torch
        print(f"Loading YOLOv11 model from: {model_path}")
        
        try:
            self.device = 'cuda:0' if torch.cuda.is_available() else 'cpu'
            if self.device != 'cpu':
                print(f" >> SUCCESS: GPU Detected: {torch.cuda.get_device_name(0)}")
            else:
                print(" >> WARNING: No GPU detected. Running on CPU.")

            # Load standard Ultralytics YOLO model
            self.model = YOLO(model_path)
            self.model.to(self.device)
            
            # Get class names dictionary
            self.model_names = self.model.names
            print(f"Model loaded successfully. Classes: {self.model_names}")
            
        except Exception as e:
            print(f"Error loading model: {e}")
            import traceback
            traceback.print_exc()
            sys.exit(1)

    def load_session(self, vp, nl, n, out):
        self.start_point = self.metadata['start_point']
        self.custom_trial = self.metadata['custom_trial']
        self.rat = self.metadata['rat']
        self.date = self.metadata['date']
        self.num_trials = self.metadata['num_trials']
        
        # --- NEW DYNAMIC LISTS ---
        self.start_nodes = self.metadata['start_nodes_list']
        self.goal_nodes = self.metadata['goal_nodes_list']
        self.trial_types = self.metadata['trial_types_list']
        # -------------------------
        
        self.special_trials = self.metadata['special_trials_list']
        self.special_start_seconds = self.metadata.get('special_start_seconds', {}) or {}
        if self.special_start_seconds:
            print("Special trial schedule (trial_num → session seconds):")
            for t_num, t_secs in sorted(self.special_start_seconds.items()):
                mm = int(t_secs // 60)
                ss = t_secs - mm * 60
                print(f"   Trial {t_num} → {mm:02d}:{ss:05.2f}")
        self.did_not_reach_list = self.metadata.get('did_not_reach_list', [])
        self.xlsx_src_path = self.metadata.get('xlsx_src_path', None)
        self.repeat = self.metadata['repeat']
        self.day_num = self.metadata['day']
        self.session_num = self.metadata['session']
        
        self.status_message = ""
        self.message_end_time = 0 
        self.all_researchers = []
        
        self.node_list = str(nl)
        self.cap = cv2.VideoCapture(str(vp))
        self.start_trial = True 
        self.end_session = False 
        self.check = False 
        self.record_detections = False 
        self.goal_location = None
        self.reached = False
        self.frame = None
        self.disp_frame = None
        self.pos_centroid = None 
        self.center_researcher = None
        
        # --- TIMER FIXES ---
        self.last_trial_end_time = -1e9  
        self.last_trial_start_time_ms = -1e9  # Added
        self.lockout_duration_ms = 10 * 60 * 1000  
        
        self.last_rat_pos = None
        self.last_researcher_pos = None
        self.prev_frame_gray = None
        self.motion_skip_threshold = 500  # changed pixels below this → skip YOLO

        if self.start_point is None:
           self.trial_num = 1
        else:
           self.trial_num = int(self.custom_trial)
        self.counter = 0 
        self.count_rat = 0

        self.start_time = 0 
        self.converted_time = "00:00:00.000"  
        
        self.normal_trial = False
        self.NGL = False
        self.probe = False
        self.probe_researcher_signalled = False
        self.start_node_delay_until = 0
        self.unnormal_intervals = self.metadata.get('unnormal_intervals', {})
        self._last_end_reason = "n/a"
        self._last_end_frame_time = -1e9
        self.trial_delays = []       # list of (trial_num, delay_seconds)
        self.trial_speed_stats = []  # list of (trial_num, avg_speed, avg_between_node_speed)
        self.trial_times = []        # list of (trial_num, start_ts, end_ts)
        self.current_trial_start_ts = ''

        self.goal_residence_timer = 0.0
        self.centroid_list = deque(maxlen=500)
        self.node_pos = []
        self.time_points = []
        self.node_id = [] 
        self.saved_nodes = []
        self.saved_velocities = []
        self.summary_trial = []
        self.store_fps = [] 
        self.locked_to_head = False   
        self.start_node_center = None
        self.covering_start_node = False
        self.cover_required_time = 10
        self.start_node_radius = 20
        self.goal_node_radius = 25
        self.save = '{}/{}_{}'.format(out, str(self.date), 'Rat' + self.rat + '.txt') 
        
        self.codec = cv2.VideoWriter_fourcc(*'mp4v')
        self.save_video = '{}/{}_{}.mp4'.format(out, str(self.date), 'Rat' + self.rat) 
        self.vid_fps = int(self.cap.get(cv2.CAP_PROP_FPS))
        
        self.out = ThreadedVideoWriter('{}'.format(self.save_video), self.codec, self.vid_fps, (1176, 712))
        
        self.researcher_goal_timer = 0.0
        self.pickup_timer = 0.0
        self.last_detection_boxes = []  # cached boxes for skipped frames

    def run_vid(self):
        print('\nStarting video processing (Live Stream Enabled).....\n')
        
        # --- GUI SETUP ---
        window_name = f"Tracker - Rat {self.rat}"
        cv2.namedWindow(window_name, cv2.WINDOW_NORMAL) 
        cv2.resizeWindow(window_name, 1176, 712) 
        # -----------------

        with open(self.save, 'w') as file:
            file.write(f"Rat number: {self.rat} , Date: {self.date} \n")
        self.Start_Time = time.time()
        
        total_frames = int(self.cap.get(cv2.CAP_PROP_FRAME_COUNT))
        frame_index = 0
        
        if self.start_point is not None:
            frame_index = int(float(self.start_point) * self.vid_fps)
            self.cap.set(cv2.CAP_PROP_POS_FRAMES, frame_index)
        
        pbar = tqdm(total=total_frames - frame_index, unit='frames', desc='Processing', ncols=100)

        while True:
            success, self.frame = self.cap.read()
            if not success:
                if not self.end_session:
                    self.calculate_velocity(self.time_points)
                    self.save_to_file(self.save)
                break

            self.frame_time = self.cap.get(cv2.CAP_PROP_POS_MSEC)
            self.converted_time = convert_milli(int(self.frame_time))
            
            frame_itr = frame_index 
            
            pbar.update(1)

            self.disp_frame = cv2.resize(self.frame, (1176, 712))
            
            self.t1 = time.time()
            self.cnn(self.disp_frame)
            self.check_special_schedule()
            self.annotate_frame(self.disp_frame)
            
            self.out.write(self.disp_frame)
            
            # --- SHOW VIDEO WINDOW (STREAM) ---
            cv2.imshow(window_name, self.disp_frame)
            
            k = cv2.waitKey(1) & 0xFF
            if k == ord('q'):
                print("\nUser interrupted execution via Window (Pressed 'q').")
                break
            # ----------------------------------
            
            # Use `is not None` (not truthy) because pos_centroid / Researcher
            # can be a numpy array; truthy check raises ValueError on multi-
            # element arrays which would kill the loop and leave an empty CSV.
            rat_x = self.pos_centroid[0] if self.pos_centroid is not None else np.nan
            rat_y = self.pos_centroid[1] if self.pos_centroid is not None else np.nan

            res_x = self.Researcher[0] if self.Researcher is not None else np.nan
            res_y = self.Researcher[1] if self.Researcher is not None else np.nan

            jp_s_x, jp_s_y = np.nan, np.nan
            jp_l_x, jp_l_y = np.nan, np.nan
            if self.record_detections:
                trial_num = self.trial_num
            else:
                trial_num = np.nan
            self.frame_data_log.append({
                'Frame_Index': frame_itr,
                'Trial_Num': trial_num,
                'Rat_X': rat_x,
                'Rat_Y': rat_y,
                'Researcher_X': res_x,
                'Researcher_Y': res_y,
                'JP_S_X': jp_s_x,
                'JP_S_Y': jp_s_y,
                'JP_L_X': jp_l_x,
                'JP_L_Y': jp_l_y
            })

            if self.record_detections:
                ts_val = self.sync_ts_dict.get("Corrected Time Stamp", {}).get(frame_itr, "N/A")
                if self.saved_nodes:
                    self.logger.info(
                        f'{self.converted_time} {ts_val} : The rat position is: {self.pos_centroid} @ {self.saved_nodes[-1]}')
                else:
                    self.logger.info(
                        f'{self.converted_time} {ts_val} : The rat position is: {self.pos_centroid}')

            if self.end_session:
                break
            
            frame_index += 1    

        pbar.close()
        
        self.export_tracking_data()
        self.post_process_xlsx()

        end = time.time()
        hours, rem = divmod(end - self.Start_Time, 3600)
        minutes, seconds = divmod(rem, 60)
        print("\nTracking process finished in: {:0>2}:{:0>2}:{:05.2f}".format(int(hours), int(minutes), seconds))
        
        self.cap.release()
        self.out.release() 
        
        cv2.destroyAllWindows()

    def export_tracking_data(self):
        print("\n>> Compiling tracking data to CSV...")

        df_tracking = pd.DataFrame(self.frame_data_log)
        print(f"   frame_data_log rows: {len(df_tracking)}")

        if df_tracking.empty:
            print("   WARNING: frame_data_log is empty — the main loop never recorded any frame.")
            print("            CSV will contain only the header row.")

        if not df_tracking.empty:
            df_tracking['Frame_Index'] = df_tracking['Frame_Index'].astype(int)

        if self.ts_file_loaded:
            try:
                # Pick the right inner dict. sync_ts_dict is shaped {col: {idx: val}}.
                # Prefer 'Corrected Time Stamp', else fall back to the column the
                # rest of the code is using (self.ts_column_name), else any column.
                ts_col_candidates = ['Corrected Time Stamp', getattr(self, 'ts_column_name', None)]
                raw_ts_data = None
                for c in ts_col_candidates:
                    if c and isinstance(self.sync_ts_dict, dict) and c in self.sync_ts_dict:
                        raw_ts_data = self.sync_ts_dict[c]
                        print(f"   Using timestamp column: {c}")
                        break
                if raw_ts_data is None and isinstance(self.sync_ts_dict, dict) and self.sync_ts_dict:
                    # Last resort: grab the first sub-dict
                    first_key = next(iter(self.sync_ts_dict))
                    candidate = self.sync_ts_dict[first_key]
                    if isinstance(candidate, dict):
                        raw_ts_data = candidate
                        print(f"   Falling back to timestamp column: {first_key}")

                if not isinstance(raw_ts_data, dict) or not raw_ts_data:
                    raise ValueError("No usable timestamp column found in sync_ts_dict")

                df_master = pd.DataFrame.from_dict(raw_ts_data, orient='index', columns=['Timestamp'])
                df_master.index.name = 'Frame_Index'
                df_master.index = df_master.index.astype(int)
                df_master.sort_index(inplace=True)
                # Promote Frame_Index from index to column so `merge(on=...)` works
                # consistently across pandas versions.
                df_master = df_master.reset_index()

                df_final = pd.merge(df_master, df_tracking, on='Frame_Index', how='left')

                if 'Timestamp_y' in df_final.columns:
                    df_final.rename(columns={'Timestamp_x': 'Timestamp'}, inplace=True)
                    df_final.drop(columns=['Timestamp_y'], inplace=True)

                df_tracking = df_final
                print(f"   After merge: {len(df_tracking)} rows, columns: {list(df_tracking.columns)}")

            except Exception as e:
                print(f"   Warning: Merge failed, saving partial data only. Error: {e}")

        cols = ['Frame_Index', 'Timestamp', 'Trial_Num', 'Rat_X', 'Rat_Y',
                'Researcher_X', 'Researcher_Y', 'JP_S_X', 'JP_S_Y', 'JP_L_X', 'JP_L_Y']

        cols = [c for c in cols if c in df_tracking.columns]
        df_tracking = df_tracking[cols]

        filename = f"{self.date}_Rat{self.rat}_Coordinates_Full.csv"
        save_path = os.path.join(self.out_path, filename)

        df_tracking.to_csv(save_path, index=False)
        print(f">> Full coordinate data saved to: {save_path}  ({len(df_tracking)} rows, {len(df_tracking.columns)} cols)")

    def check_special_schedule(self):
        """If a time-locked special trial's unlock time has arrived while an
        earlier trial is still active, force-end the earlier trial so the
        special trial's start_node becomes triggerable on the next frame."""
        if not self.special_start_seconds or not self.record_detections:
            return
        elapsed_s = self.frame_time / 1000.0
        # Find the earliest scheduled trial that is still in the future of
        # the current active trial and whose unlock time has been reached.
        for sp_trial_num, sp_unlock_s in self.special_start_seconds.items():
            if sp_trial_num <= self.trial_num:
                continue  # past or current
            if elapsed_s >= sp_unlock_s:
                print(f"\n[SCHEDULE] Trial {sp_trial_num} unlock time {sp_unlock_s:.2f}s "
                      f"reached at session {elapsed_s:.2f}s — force-ending active trial {self.trial_num}.")
                self.end_trial(reason="forced by special trial schedule")
                return

    def find_start(self, center_rat):
        # Time-locked special trial: its start_node won't trigger before
        # the scheduled session time. The trial is "hidden" until then.
        if self.trial_num in self.special_start_seconds:
            elapsed_s = self.frame_time / 1000.0
            if elapsed_s < self.special_start_seconds[self.trial_num]:
                return

        node = self.start_nodes_locations[self.counter]
        self.locked_to_head = False
        if points_dist(center_rat, node) < 60:
            self.logger.info('Recording Trial {}'.format(self.trial_num))
            
            # --- RECORD TRIAL START TIME ---
            self.last_trial_start_time_ms = self.frame_time
            curr_idx = int(self.cap.get(cv2.CAP_PROP_POS_FRAMES)) - 1
            self.current_trial_start_ts = self.sync_ts_dict.get(self.ts_column_name, {}).get(curr_idx, '')
            # -------------------------------
            
            current_trial_type = int(self.trial_types[self.counter])
            self.goal_location = self.goal_locations[self.counter]
            self.current_goal_name = self.goal_nodes[self.counter]
            
            if self.trial_num == 1 and current_trial_type != 1:
                self.start_time = (self.frame_time / (1000 * 60)) % 60
                if current_trial_type == 3:
                    self.probe = True
                if current_trial_type == 2:
                    self.NGL = True
                    
            if current_trial_type in [4, 5, 6]:
                self.NGL = True
                self.start_time = (self.frame_time / (1000 * 60)) % 60
                    
            if not self.probe and not self.NGL:
                self.normal_trial = True

            self.node_pos = []
            self.centroid_list = []
            self.time_points = []
            self.summary_trial = []
            self.saved_nodes = []
            self.node_id = [] 
            self.saved_velocities = []
            self.record_detections = True

            self.researcher_goal_timer = 0.0
            self.pickup_timer = 0.0
            
            self.pos_centroid = node
            self.centroid_list.append(self.pos_centroid)
            self.start_trial = False
            
    def check_immunity(self):
        if self.trial_num in self.unnormal_intervals:
            start_block, end_block = self.unnormal_intervals[self.trial_num]
            current_abs_minutes = (self.frame_time / (1000 * 60)) % 60
            if start_block <= current_abs_minutes <= end_block:
                return True
        return False
    
    def closest_researcher_to(self, point):
        """Return the closest researcher position to the given point, or None."""
        if not self.all_researchers:
            return None
        return min(self.all_researchers, key=lambda r: points_dist(r, point))

    def _load_maze_roi(self, _unused):
        """Load src/tools/maze_roi.txt (committed to repo, shared by all users)."""
        roi_path = Path(__file__).parent / "tools" / "maze_roi.txt"
        if roi_path.exists():
            try:
                points = []
                for line in roi_path.read_text().splitlines():
                    line = line.strip()
                    if not line or line.startswith("#"):
                        continue
                    x, y = line.split(",")
                    points.append((int(x), int(y)))
                self.maze_roi = np.array(points, dtype=np.int32)
                print(f"Maze ROI loaded: {len(points)} vertices from {roi_path}")
            except Exception as e:
                print(f"Warning: could not load maze_roi.txt: {e}")
                self.maze_roi = None
        else:
            self.maze_roi = None
            print("No maze_roi.txt found — rat detection not spatially restricted.")

    @staticmethod
    def _box_overlap_ratio(box_a, box_b):
        """Fraction of box_a's area that is covered by box_b (0.0–1.0)."""
        ax1, ay1, ax2, ay2 = box_a
        bx1, by1, bx2, by2 = box_b
        ix1, iy1 = max(ax1, bx1), max(ay1, by1)
        ix2, iy2 = min(ax2, bx2), min(ay2, by2)
        if ix2 <= ix1 or iy2 <= iy1:
            return 0.0
        inter = (ix2 - ix1) * (iy2 - iy1)
        area_a = max((ax2 - ax1) * (ay2 - ay1), 1)
        return inter / area_a

    def compute_motion(self, frame):
        """Return True if enough pixels changed since the last frame, False otherwise.
        Always updates self.prev_frame_gray so comparisons stay current even on skipped frames."""
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        gray = cv2.GaussianBlur(gray, (5, 5), 0)
        prev = self.prev_frame_gray
        self.prev_frame_gray = gray

        if prev is None:
            return True  # first frame — always detect

        diff = cv2.absdiff(gray, prev)
        _, thresh = cv2.threshold(diff, 20, 255, cv2.THRESH_BINARY)
        thresh = cv2.dilate(thresh, None, iterations=3)
        return cv2.countNonZero(thresh) >= self.motion_skip_threshold

    def cnn(self, frame):
        has_motion = self.compute_motion(frame)

        self.Rat = None
        self.Researcher = None
        self.all_researchers = []

        rat_candidates = []
        researcher_candidates = []

        detected_rat_body_this_frame = False

        if has_motion:
            results = self.model(frame, conf=0.7, verbose=False, imgsz=1280)
            current_boxes = []
            for r in results:
                boxes = r.boxes
                for box in boxes:
                    x1, y1, x2, y2 = box.xyxy[0].cpu().numpy().astype(int)
                    confidence = float(box.conf[0])
                    cls_id = int(box.cls[0])
                    label = self.model_names[cls_id]
                    centroid = (int((x1 + x2) / 2), int((y1 + y2) / 2))
                    current_boxes.append((x1, y1, x2, y2, label, confidence, cls_id))

                    if label == 'rat':
                        rat_candidates.append((confidence, centroid, x1, y1, x2, y2))
                        detected_rat_body_this_frame = True
                    elif label == 'researcher':
                        researcher_candidates.append((confidence, centroid, x1, y1, x2, y2))

            # Only update cache when YOLO actually found something;
            # keeping stale boxes on missed frames prevents flash
            if current_boxes:
                self.last_detection_boxes = current_boxes

        # Always redraw last known boxes so display doesn't flash on skipped frames
        for x1, y1, x2, y2, label, confidence, cls_id in self.last_detection_boxes:
            color = colors[cls_id % len(colors)]
            cv2.rectangle(self.disp_frame, (x1, y1), (x2, y2), color, 2)
            cv2.putText(self.disp_frame, f"{label} {confidence:.2f}",
                        (x1, y1 + 20), font, 1, (255, 255, 255), 1)

        # --- RESEARCHER SELECTION: store ALL positions ---
        researcher_boxes = [(rx1, ry1, rx2, ry2) for _, _, rx1, ry1, rx2, ry2 in researcher_candidates]
        if researcher_candidates:
            self.all_researchers = [pos for _, pos, *_ in researcher_candidates]

        # --- RAT SELECTION: highest-confidence body not overlapping or inside a researcher ---
        RAT_OVERLAP_THRESHOLD = 0.1  # reject if >10 % of rat box is covered by a researcher box
        if rat_candidates:
            rat_candidates.sort(key=lambda x: x[0], reverse=True)
            for conf, centroid, rx1, ry1, rx2, ry2 in rat_candidates:
                rat_box = (rx1, ry1, rx2, ry2)
                cx, cy = centroid
                rejected = False
                for rb in researcher_boxes:
                    # centroid inside researcher box → rat is being held
                    rbx1, rby1, rbx2, rby2 = rb
                    if rbx1 <= cx <= rbx2 and rby1 <= cy <= rby2:
                        rejected = True
                        break
                    # significant box overlap
                    if self._box_overlap_ratio(rat_box, rb) > RAT_OVERLAP_THRESHOLD:
                        rejected = True
                        break
                if not rejected and self.maze_roi is not None:
                    # reject rat whose centroid is outside the maze polygon
                    if cv2.pointPolygonTest(self.maze_roi, (float(cx), float(cy)), False) < 0:
                        rejected = True

                if not rejected:
                    self.Rat = centroid
                    break

        # --- CACHE LAST KNOWN POSITIONS ---
        if self.Rat is not None:
            self.last_rat_pos = self.Rat
        if self.all_researchers:
            self.last_researcher_pos = self.all_researchers[0]

        # --- ACTIVE RAT POS (must be set before anything else uses it) ---
        active_rat_pos = self.Rat if self.Rat is not None else self.last_rat_pos

        # --- Set self.Researcher to closest to rat (for logging / frame_data_log) ---
        if self.all_researchers and active_rat_pos:
            self.Researcher = self.closest_researcher_to(active_rat_pos)
        elif self.all_researchers:
            self.Researcher = self.all_researchers[0]
        else:
            self.Researcher = None

        # --- UNNORMAL INTERVAL FORCE-END ---
        if not self.start_trial and not self.end_session and self.trial_num in self.unnormal_intervals:
            _, end_block_abs = self.unnormal_intervals[self.trial_num]
            current_abs_minutes = (self.frame_time / (1000 * 60)) % 60

            if current_abs_minutes >= end_block_abs:
                self.normal_trial = False
                self.NGL = False
                self.probe = False
                self.probe_researcher_signalled = False
                self.reached = False
                self.end_trial(reason="unnormal interval timeout")
                self.start_trial = True
                self.check = False
                return

        # --- RESEARCHER TRIGGER: start next trial (closest researcher to RAT) ---
        if active_rat_pos and not self.record_detections:
            closest_to_rat = self.closest_researcher_to(active_rat_pos)

            if closest_to_rat is not None:
                dist = points_dist(active_rat_pos, closest_to_rat)

                if self.counter > 0 and (self.counter - 1) < len(self.trial_types):
                    previous_trial_type = int(self.trial_types[self.counter - 1])
                else:
                    previous_trial_type = 1

                is_special_lockout = previous_trial_type in [4, 5, 6]

                time_since_trial_start = self.frame_time - getattr(self, 'last_trial_start_time_ms', -1e9)

                can_trigger = True
                if is_special_lockout:
                    if time_since_trial_start < self.lockout_duration_ms:
                        can_trigger = False
                        remaining_sec = int((self.lockout_duration_ms - time_since_trial_start) / 1000)
                        cv2.putText(self.disp_frame, f"LOCKOUT: {remaining_sec}s", (60, 110),
                                    font, 1, (0, 0, 255), 2)
                    else:
                        cv2.putText(self.disp_frame, "READY: Researcher can start trial", (60, 110),
                                    font, 1, (0, 255, 0), 2)

                if (not self.start_trial and not self.end_session and
                    not self.record_detections and dist <= 300 and can_trigger):

                    print(f">>> Lockout finished/not required. Starting Trial {self.trial_num}")
                    self.start_trial = True
                    self.check = False

        # --- RAT DETECTION + TRIAL LOGIC ---
        if active_rat_pos:
            if self.start_trial:
                self.find_start(active_rat_pos)

            if self.record_detections and detected_rat_body_this_frame:
                self.count_rat += 1

                self.object_detection(rat=active_rat_pos)

                # Re-check record_detections AFTER object_detection,
                # because object_detection -> end_trial() may have set it to False.
                if self.record_detections and self.goal_location is not None:
                    # Use closest researcher to the GOAL for 10s force-end
                    closest_to_goal = self.closest_researcher_to(self.goal_location)

                    if closest_to_goal is not None:
                        dist_to_goal = points_dist(closest_to_goal, self.goal_location)

                        if dist_to_goal <= 50:
                            self.researcher_goal_timer += (1.0 / self.vid_fps)

                            if self.researcher_goal_timer >= 10.0:
                                self.normal_trial = False
                                self.NGL = False
                                self.probe = False
                                self.probe_researcher_signalled = False
                                self.end_trial(reason="researcher at goal 10s")
                                self.researcher_goal_timer = 0.0
                        else:
                            self.researcher_goal_timer = 0.0

        # --- RESEARCHER AT GOAL: 30-second force-end (closest to GOAL) ---
        if self.record_detections and self.goal_location is not None:
            closest_to_goal = self.closest_researcher_to(self.goal_location)

            if closest_to_goal is not None:
                dist_to_goal = points_dist(closest_to_goal, self.goal_location)

                if dist_to_goal <= 160:

                    allow_end = True

                    if self.probe:
                        current_min = (self.frame_time / (1000 * 60)) % 60
                        duration = current_min - self.start_time
                        if duration < 0: duration += 60
                        if duration < 2.0:
                            allow_end = False

                    if self.check_immunity():
                        allow_end = False

                    if allow_end:
                        self.researcher_goal_timer += (1.0 / self.vid_fps)

                        if self.researcher_goal_timer >= 30.0:
                            self.normal_trial = False
                            self.NGL = False
                            self.probe = False
                            self.probe_researcher_signalled = False

                            self.end_trial(reason="researcher at goal 30s")
                            self.researcher_goal_timer = 0.0
                    else:
                        self.researcher_goal_timer = 0.0

        # --- RESEARCHER COVERING START NODE (closest to START NODE) ---
        researcher_covers_start = False
        if (not self.start_trial and not self.record_detections and
            not self.end_session and self.counter < len(self.start_nodes_locations)):

            self.start_node_center = self.start_nodes_locations[self.counter]
            closest_to_start = self.closest_researcher_to(self.start_node_center)

            if closest_to_start is not None:
                dist_to_start = points_dist(closest_to_start, self.start_node_center)
                if dist_to_start <= 40:
                    researcher_covers_start = True

            if researcher_covers_start:
                if not self.covering_start_node:
                    self.covering_start_node = True
                    self.cover_start_timer = 0.0

                self.cover_start_timer += self.frame_time

                if self.cover_start_timer >= self.cover_required_time:
                    self.start_trial = True
                    self.check = False
                    self.covering_start_node = False
                    self.cover_start_timer = 0.0
            else:
                if self.covering_start_node:
                    self.covering_start_node = False
                    self.cover_start_timer = 0.0

    def object_detection(self, rat):
        self.pos_centroid = rat
        self.centroid_list.append(self.pos_centroid)

        is_immune = self.check_immunity()
        is_did_not_reach = (self.counter < len(self.did_not_reach_list) and
                            self.did_not_reach_list[self.counter] == 1)

        if self.NGL:
            minutes = self.timer(start=self.start_time)
            if not self.reached:
                if points_dist(self.pos_centroid, self.goal_location) <= 20:
                    self.reached = True
            if minutes >= 10:
                print('\n\n >>> Ten minute passed... Goal location reached:', self.reached)
                if self.reached:
                    if not is_immune:
                        print('\n\n >>> End New Goal Location Trial - timeout', self.trial_num, ' out of ',
                            self.num_trials)
                        self.NGL = False
                        self.reached = False
                        self.end_trial(reason="NGL 10min timeout")

        if self.probe:
            minutes = self.timer(start=self.start_time)
            if minutes > 2:
                if not self.probe_researcher_signalled:
                    closest_to_goal = self.closest_researcher_to(self.goal_location)
                    if (closest_to_goal is not None and
                            points_dist(closest_to_goal, self.goal_location) <= 80):
                        self.probe_researcher_signalled = True
                else:
                    if points_dist(self.pos_centroid, self.goal_location) <= self.goal_node_radius:
                        if not is_immune:
                            self.probe = False
                            self.probe_researcher_signalled = False
                            self.end_trial(reason="probe complete")

        if self.normal_trial:
            if not is_did_not_reach:
                if points_dist(self.pos_centroid, self.goal_location) <= self.goal_node_radius:
                    if not is_immune:
                        self.normal_trial = False
                        self.end_trial(reason="normal reached goal")
            else:
                # "Did Not Reach" end logic: trial ends when rat is picked up by researcher
                # Use closest researcher to the RAT for pickup detection
                closest_to_rat = self.closest_researcher_to(self.pos_centroid)
                if closest_to_rat is not None:
                    dist_to_researcher = points_dist(self.pos_centroid, closest_to_rat)
                    if dist_to_researcher <= 60:
                        self.pickup_timer += (1.0 / self.vid_fps)
                        if self.pickup_timer >= 1.0:
                            print(f'\n\n >>> Did Not Reach: Trial {self.trial_num} ended - rat picked up by researcher')
                            self.normal_trial = False
                            self.end_trial(reason="DNR rat picked up")
                            self.pickup_timer = 0.0
                    else:
                        self.pickup_timer = 0.0

    def object_detection(self, rat):
        self.pos_centroid = rat
        self.centroid_list.append(self.pos_centroid)

        is_immune = self.check_immunity()
        is_did_not_reach = (self.counter < len(self.did_not_reach_list) and
                            self.did_not_reach_list[self.counter] == 1)

        if self.NGL:
            minutes = self.timer(start=self.start_time)
            if not self.reached:
                if points_dist(self.pos_centroid, self.goal_location) <= 20:
                    self.reached = True
            if minutes >= 10:
                print('\n\n >>> Ten minute passed... Goal location reached:', self.reached)
                if self.reached:
                    if not is_immune:
                        print('\n\n >>> End New Goal Location Trial - timeout', self.trial_num, ' out of ',
                            self.num_trials)
                        self.NGL = False
                        self.reached = False
                        self.end_trial(reason="NGL 10min timeout")

        trial_elapsed_ms = self.frame_time - self.last_trial_start_time_ms
        researcher_trigger_allowed = trial_elapsed_ms >= 5_000

        # For all trial types except 3, 4, 5, 6: end trial when researcher is within 150px of the rat
        if researcher_trigger_allowed:
            _curr_type = int(self.trial_types[self.counter]) if self.counter < len(self.trial_types) else 1
            if _curr_type not in (3, 4, 5, 6):
                _closest_to_rat = self.closest_researcher_to(self.pos_centroid)
                if _closest_to_rat is not None:
                    _res_rat_dist = points_dist(_closest_to_rat, self.pos_centroid)
                    if _res_rat_dist <= 150:
                        print(f'\n\n >>> Trial {self.trial_num} (type {_curr_type}): researcher within 150px of rat ({_res_rat_dist:.0f}px), ending trial')
                        self.normal_trial = False
                        self.NGL = False
                        self.probe = False
                        self.probe_researcher_signalled = False
                        self.end_trial(reason="researcher near rat 150px")
                        return

        if self.probe:
            minutes = self.timer(start=self.start_time)
            _dbg_sec = int(self.frame_time / 1000)
            if minutes >= 2:
                if not self.probe_researcher_signalled:
                    closest_to_goal = self.closest_researcher_to(self.goal_location)
                    if closest_to_goal is not None:
                        dist = points_dist(closest_to_goal, self.goal_location)
                        if _dbg_sec != getattr(self, '_probe_dbg_sec', -1):
                            self._probe_dbg_sec = _dbg_sec
                            print(f"[PROBE] min={minutes} | res->goal={dist:.1f}px | threshold=600px | will_signal={dist<=600}")
                        if dist <= 600:
                            self.probe_researcher_signalled = True
                    else:
                        if _dbg_sec != getattr(self, '_probe_dbg_sec', -1):
                            self._probe_dbg_sec = _dbg_sec
                            print(f"[PROBE] min={minutes} | no researcher detected this frame")
                else:
                    if points_dist(self.pos_centroid, self.goal_location) <= self.goal_node_radius:
                        if not is_immune:
                            self.probe = False
                            self.probe_researcher_signalled = False
                            self.end_trial(reason="probe complete")
            else:
                if _dbg_sec != getattr(self, '_probe_dbg_sec', -1):
                    self._probe_dbg_sec = _dbg_sec
                    print(f"[PROBE] waiting: {minutes}min elapsed (need >= 2)")

        if self.normal_trial:
            if points_dist(self.pos_centroid, self.goal_location) <= self.goal_node_radius:
                if not is_immune:
                    self.normal_trial = False
                    self.end_trial(reason="normal reached goal")

    def end_trial(self, reason="unknown"):
        self._last_end_reason = reason
        self._last_end_frame_time = self.frame_time
        _delay_s = round((self.frame_time - self.last_trial_start_time_ms) / 1000, 2)
        self.trial_delays.append((self.trial_num, _delay_s))
        _curr_idx = int(self.cap.get(cv2.CAP_PROP_POS_FRAMES)) - 1
        _end_ts = self.sync_ts_dict.get(self.ts_column_name, {}).get(_curr_idx, '')
        self.trial_times.append((self.trial_num, self.current_trial_start_ts, _end_ts))
        print(f'\n[END_TRIAL] trial={self.trial_num} counter={self.counter} reason="{reason}" '
              f'frame_time={self.frame_time/1000:.2f}s '
              f'normal={self.normal_trial} NGL={self.NGL} probe={self.probe} immune={self.check_immunity()}')

        # Decide BEFORE we overwrite pos_centroid whether the rat actually
        # reached the goal this trial. Two sources of truth:
        #   (a) reason string is one of the "rat reached goal" branches
        #   (b) rat's real position right now is within goal_node_radius of goal
        # (b) catches cases where the researcher-proximity check fires before
        # the goal-reached check (researcher walks in to pick rat up at goal).
        SUCCESS_REASONS = ("normal reached goal", "probe complete", "NGL 10min timeout")
        rat_reached_goal = reason in SUCCESS_REASONS
        dist_to_goal_at_end = None
        if self.goal_location is not None and self.pos_centroid is not None:
            dist_to_goal_at_end = points_dist(self.pos_centroid, self.goal_location)
            if not rat_reached_goal and dist_to_goal_at_end <= self.goal_node_radius:
                rat_reached_goal = True

        self.pos_centroid = self.goal_location
        self.centroid_list.append(self.pos_centroid)

        # If the rat reached the goal, force-record the goal node in the path.
        # Without this, the node-detection loop in annotate_frame is skipped
        # (record_detections is about to be False), so the goal node often
        # never appears in saved_nodes even though the rat reached it.
        goal_injected = False
        last_saved_before = self.saved_nodes[-1] if self.saved_nodes else None
        if rat_reached_goal:
            goal_name_str = (str(self.current_goal_name)
                             if getattr(self, 'current_goal_name', None) is not None else None)
            if goal_name_str is not None and (not self.saved_nodes or self.saved_nodes[-1] != goal_name_str):
                _curr_idx = int(self.cap.get(cv2.CAP_PROP_POS_FRAMES)) - 1
                sync_time = self.sync_ts_dict.get(self.ts_column_name, {}).get(_curr_idx, self.converted_time)
                self.saved_nodes.append(goal_name_str)
                self.node_pos.append(self.goal_location)
                self.time_points.append([sync_time, goal_name_str])
                goal_injected = True

        # Capture debug info for save_to_file to write into the .txt summary.
        self._last_trial_debug = {
            'reason': reason,
            'dist_to_goal_px': (round(dist_to_goal_at_end, 1) if dist_to_goal_at_end is not None else None),
            'goal_radius_px': self.goal_node_radius,
            'rat_reached_goal': rat_reached_goal,
            'goal_name': (str(self.current_goal_name) if getattr(self, 'current_goal_name', None) is not None else None),
            'last_node_before_inject': last_saved_before,
            'goal_injected': goal_injected,
        }

        self.record_detections = False  # disable detections before annotate_frame so post-end frames don't get injected
        self.annotate_frame(self.disp_frame)

        self.calculate_velocity(self.time_points)

        if self.summary_trial:
            total_dist = sum(seg[3] for seg in self.summary_trial)
            total_time_s = sum(seg[2] for seg in self.summary_trial)
            avg_speed = round(total_dist / total_time_s, 3) if total_time_s > 0 else ''
        else:
            avg_speed = ''
        avg_between_node = round(sum(self.saved_velocities) / len(self.saved_velocities), 3) if self.saved_velocities else ''
        self.trial_speed_stats.append((self.trial_num, avg_speed, avg_between_node))

        self.save_to_file(self.save)
        self.last_trial_end_time = self.frame_time

        self.counter += 1

        if self.counter < int(self.num_trials):
            self.trial_num += 1
        else:
            self.end_session = True

        self.count_rat = 0

        print(f'[END_TRIAL] → next trial_num={self.trial_num} counter={self.counter} end_session={self.end_session}')


    def timer(self, start):
        end = (self.frame_time / (1000 * 60)) % 60
        duration = end - start
        if duration < 0:
            duration = duration + 60
        return int(duration)

    def calculate_velocity(self, time_points):
        bridges = {('124', '201'): 0.60,
                   ('121', '302'): 1.72,
                   ('223', '404'): 1.69,
                   ('324', '401'): 0.60,
                   ('305', '220'): 0.60}
        
        if len(time_points) > 2:
            for i in range(0, len(time_points) - 1):
                start_node = time_points[i][1]
                j = i + 1
                end_node = time_points[j][1]
                
                try:
                    # Logic to handle BOTH old string timestamps and new float timestamps
                    t1 = time_points[i][0]
                    t2 = time_points[j][0]
                    
                    if isinstance(t1, str) and ":" in t1:
                        # Old behavior for backward compatibility
                        format = '%H:%M:%S.%f'
                        st = datetime.strptime(t1, format)
                        et = datetime.strptime(t2, format)
                        difference = (et - st).total_seconds()
                    else:
                        # New behavior: Directly subtract the "Seconds From Creation"
                        difference = float(t2) - float(t1)

                    # Calculate distance
                    if (start_node, end_node) in bridges or (end_node, start_node) in bridges:
                        lenght = bridges.get((start_node, end_node), bridges.get((end_node, start_node)))
                    else:
                        lenght = 0.30 

                    # Calculate speed
                    if difference > 0:
                        speed = round(float(lenght) / float(difference), 3)
                    else:
                        speed = 0

                    self.summary_trial.append(
                        [(start_node, end_node), (t1, t2), round(difference, 3), lenght, speed])
                    self.saved_velocities.append(speed)

                except Exception as e:
                    print(f"Error calculating velocity at nodes {start_node}->{end_node}: {e}")
                    continue

    @staticmethod
    def annotate_node(frame, point, node, t):
        if t == 1:
            cv2.circle(frame, point, 20, color=(0, 255, 0), thickness=2)
            cv2.putText(frame, str(node), (point[0] - 16, point[1]),
                        fontScale=0.5, fontFace=FONT, color=(0, 255, 0), thickness=1,
                        lineType=cv2.LINE_AA)
            cv2.putText(frame, 'Start', (point[0] - 16, point[1] - 22),
                        fontScale=0.5, fontFace=FONT, color=(0, 255, 0), thickness=1,
                        lineType=cv2.LINE_AA)

        if t == 2:
            cv2.circle(frame, point, 20, color=(20, 110, 245), thickness=1)
            cv2.putText(frame, str(node), (point[0] - 16, point[1]),
                        fontScale=0.5, fontFace=FONT, color=(0, 69, 255), thickness=1,
                        lineType=cv2.LINE_AA)
        if t == 3:
            cv2.circle(frame, point, 20, color=(0, 0, 250), thickness=2)
            cv2.putText(frame, str(node), (point[0] - 16, point[1]),
                        fontScale=0.5, fontFace=FONT, color=(0, 0, 255), thickness=1,
                        lineType=cv2.LINE_AA)
            cv2.putText(frame, 'End', (point[0] - 16, point[1] - 22),
                        fontScale=0.5, fontFace=FONT, color=(0, 0, 255), thickness=1,
                        lineType=cv2.LINE_AA)

    def annotate_frame(self, frame):
        nodes_dict = self.nodes_dict 
        
        cv2.putText(frame, str(self.converted_time), (970, 670),
                    fontFace=FONT, fontScale=0.75, color=(240, 240, 240), thickness=1)
        
        time_diff = time.time() - self.t1
        fps = 1.0 / max(time_diff, 0.001) 
        
        self.store_fps.append(fps)
        cv2.putText(frame, "FPS: {:.2f}".format(fps), (970, 650), fontFace=FONT, fontScale=0.75, color=(240, 240, 240),
                    thickness=1)
        
        if self.counter < len(self.goal_locations):
            active_goal_loc = self.goal_locations[self.counter]
            active_goal_name = self.goal_nodes[self.counter]
            if active_goal_loc is not None:
                self.annotate_node(frame, point=active_goal_loc, node=active_goal_name, t=3)
        
        if self.start_trial and self.counter < len(self.start_nodes):
            cv2.putText(frame, f'Next trial: {self.trial_num}', (60, 60),
                        fontFace=FONT, fontScale=0.75, color=(255, 255, 255), thickness=1)

            _type_names = {1: "Normal", 2: "NGL", 3: "Probe", 4: "NGL-Sp4", 5: "NGL-Sp5", 6: "NGL-Sp6"}
            if self.counter < len(self.trial_types):
                _next_type = int(self.trial_types[self.counter])
                cv2.putText(frame, f'Type: {_type_names.get(_next_type, f"Type {_next_type}")}', (60, 100),
                            fontFace=FONT, fontScale=0.65, color=(180, 220, 255), thickness=1)
            _next_goal = str(self.goal_nodes[self.counter]) if self.counter < len(self.goal_nodes) else "?"
            cv2.putText(frame, f'Goal: {_next_goal}', (60, 118),
                        fontFace=FONT, fontScale=0.65, color=(180, 255, 180), thickness=1)

            if self.frame_time < self.start_node_delay_until:
                _remaining = (self.start_node_delay_until - self.frame_time) / 1000
                cv2.putText(frame, f'Start node in {_remaining:.1f}s...', (60, 80),
                            fontFace=FONT, fontScale=0.75, color=(255, 200, 80), thickness=1)
            else:
                cv2.putText(frame, 'Waiting start new trial...', (60, 80),
                            fontFace=FONT, fontScale=0.75, color=(255, 255, 255), thickness=1)
                start_pos = self.start_nodes_locations[self.counter]
                start_node_name = self.start_nodes[self.counter]
                self.annotate_node(frame, point=start_pos, node=start_node_name, t=1)

        # --- BETWEEN-TRIALS DEBUG OVERLAY ---
        if not self.record_detections and not self.end_session:
            # Last end reason
            _secs_since_end = (self.frame_time - self._last_end_frame_time) / 1000
            cv2.putText(frame, f'Last end: "{self._last_end_reason}" ({_secs_since_end:.1f}s ago)',
                        (60, 172), fontFace=FONT, fontScale=0.55, color=(200, 200, 80), thickness=1)

            # start_trial trigger A: researcher near rat (160px)
            _st_rat_pos = getattr(self, 'last_rat_pos', None)
            if _st_rat_pos:
                _st_closest = self.closest_researcher_to(_st_rat_pos)
                if _st_closest is not None:
                    _st_dist = points_dist(_st_closest, _st_rat_pos)
                    # lockout check
                    _prev_type = int(self.trial_types[self.counter - 1]) if self.counter > 0 and (self.counter - 1) < len(self.trial_types) else 1
                    _lockout_active = _prev_type in [1] and (self.frame_time - getattr(self, 'last_trial_start_time_ms', -1e9)) < self.lockout_duration_ms
                    _lockout_rem = max(0, self.lockout_duration_ms - (self.frame_time - getattr(self, 'last_trial_start_time_ms', -1e9))) / 1000
                    _trig_color = (0, 200, 0) if _st_dist <= 160 and not _lockout_active else (180, 180, 180)
                    _lockout_str = f' [LOCKOUT {_lockout_rem:.0f}s]' if _lockout_active else ''
                    cv2.putText(frame, f'TrigA res->rat {_st_dist:.0f}px/300{_lockout_str} | start_trial={self.start_trial}',
                                (60, 190), fontFace=FONT, fontScale=0.55, color=_trig_color, thickness=1)
                else:
                    cv2.putText(frame, 'TrigA: no researcher detected',
                                (60, 190), fontFace=FONT, fontScale=0.55, color=(180, 180, 180), thickness=1)

            # start_trial trigger B: researcher covers start node (40px)
            if self.counter < len(self.start_nodes_locations):
                _sn_center = self.start_nodes_locations[self.counter]
                _sn_closest = self.closest_researcher_to(_sn_center)
                if _sn_closest is not None:
                    _sn_dist = points_dist(_sn_closest, _sn_center)
                    _sn_color = (0, 200, 0) if _sn_dist <= 40 else (180, 180, 180)
                    _cover_t = getattr(self, 'cover_start_timer', 0)
                    cv2.putText(frame, f'TrigB res->start_node {_sn_dist:.0f}px/40 cover_t={_cover_t:.0f}ms',
                                (60, 208), fontFace=FONT, fontScale=0.55, color=_sn_color, thickness=1)
                else:
                    cv2.putText(frame, 'TrigB: no researcher at start node',
                                (60, 208), fontFace=FONT, fontScale=0.55, color=(180, 180, 180), thickness=1)

        # Inside annotate_frame, find the 'record_detections' block:
        if self.record_detections:
            # Get the sync time for the current frame
            curr_idx = int(self.cap.get(cv2.CAP_PROP_POS_FRAMES)) - 1
            sync_time = self.sync_ts_dict.get(self.ts_column_name, {}).get(curr_idx, self.converted_time)

            # Goal gets a wider detect radius (26 px) so the trial-ending check
            # (25 px on self.goal_location) is guaranteed to be preceded by the
            # goal being recorded into saved_nodes. Other nodes stay at 20 px.
            goal_key = str(getattr(self, 'current_goal_name', '')) if getattr(self, 'current_goal_name', None) is not None else None
            for node_name in nodes_dict:
                detect_radius = 26 if (goal_key is not None and str(node_name) == goal_key) else 20
                if points_dist(self.pos_centroid, nodes_dict[node_name]) <= detect_radius:
                    self.saved_nodes.append(node_name)
                    self.node_pos.append(nodes_dict[node_name])

                    # Use sync_time instead of self.converted_time
                    if len(self.time_points) == 0:
                        self.time_points.append([sync_time, node_name])
                    elif node_name != self.saved_nodes[-2]:
                        self.time_points.append([sync_time, node_name])

            cv2.putText(frame, 'Trial:' + str(self.trial_num), (60, 60),
                        fontFace=FONT, fontScale=0.75, color=(255, 255, 255), thickness=1)
            cv2.putText(frame, 'Currently writing to file...', (60, 80),
                        fontFace=FONT, fontScale=0.75, color=(255, 255, 255), thickness=1)

            _type_names = {1: "Normal", 2: "NGL", 3: "Probe", 4: "NGL-Sp4", 5: "NGL-Sp5", 6: "NGL-Sp6"}
            _curr_type = int(self.trial_types[self.counter]) if self.counter < len(self.trial_types) else -1
            cv2.putText(frame, f'Type: {_type_names.get(_curr_type, f"Type {_curr_type}")}', (60, 100),
                        fontFace=FONT, fontScale=0.65, color=(180, 220, 255), thickness=1)

            _conds = []
            if self.normal_trial: _conds.append("normal")
            if self.NGL: _conds.append("NGL")
            if self.probe: _conds.append("probe")
            if self.probe: _conds.append(f'res.signalled={"T" if self.probe_researcher_signalled else "F"}')
            if self.NGL and self.reached: _conds.append("reached_goal")
            _is_dnr = (self.counter < len(self.did_not_reach_list) and
                       self.did_not_reach_list[self.counter] == 1)
            if _is_dnr: _conds.append("DNR")
            if self.check_immunity(): _conds.append("immune")
            cv2.putText(frame, f'Conds: {", ".join(_conds) if _conds else "none"}', (60, 118),
                        fontFace=FONT, fontScale=0.65, color=(180, 220, 255), thickness=1)

            if self.pos_centroid and self.goal_location:
                _dist_to_goal = points_dist(self.pos_centroid, self.goal_location)
                cv2.putText(frame, f'Dist to goal: {_dist_to_goal:.1f}px', (60, 136),
                            fontFace=FONT, fontScale=0.65, color=(255, 200, 100), thickness=1)

            _curr_type_dbg = int(self.trial_types[self.counter]) if self.counter < len(self.trial_types) else 1
            if _curr_type_dbg not in (3, 4, 5, 6) and self.pos_centroid:
                _closest = self.closest_researcher_to(self.pos_centroid)
                if _closest is not None:
                    _res_rat_dist = points_dist(_closest, self.pos_centroid)
                    _res_rat_color = (0, 60, 255) if _res_rat_dist <= 150 else (255, 255, 255)
                    _res_rat_label = f'Res->rat {_res_rat_dist:.0f}px (thr:150) - {"ENDING" if _res_rat_dist <= 150 else "waiting"}'
                else:
                    _res_rat_color = (255, 255, 255)
                    _res_rat_label = 'Res->rat: no researcher detected'
                cv2.putText(frame, _res_rat_label, (60, 154),
                            fontFace=FONT, fontScale=0.65, color=_res_rat_color, thickness=1)

            if self.probe and self.goal_location:
                _probe_min = (self.frame_time / (1000 * 60)) % 60 - self.start_time
                if _probe_min < 0: _probe_min += 60
                _timer_int = int(_probe_min)
                _closest_res = self.closest_researcher_to(self.goal_location)
                if _closest_res:
                    _res_d = points_dist(_closest_res, self.goal_location)
                    _res_dist_str = f'{_res_d:.0f}px (thr:600)'
                else:
                    _res_dist_str = 'no res'
                _timer_label = f'{_probe_min:.2f}min [int={_timer_int}]'
                cv2.putText(frame, f'Probe: {_timer_label} | Res->goal: {_res_dist_str}', (60, 154),
                            fontFace=FONT, fontScale=0.65, color=(255, 140, 80), thickness=1)

            cv2.putText(frame, "Rat Count: " + str(self.count_rat), (40, 172),
                        fontFace=FONT, fontScale=0.65, color=(255, 255, 255), thickness=1)

            if len(self.centroid_list) >= 2:
                for i in range(1, len(self.centroid_list)):
                    cv2.line(frame, self.centroid_list[i], self.centroid_list[i - 1],
                             color=(255, 0, 60), thickness=1)
            cv2.line(frame, (self.pos_centroid[0] - 5, self.pos_centroid[1]),
                     (self.pos_centroid[0] + 5, self.pos_centroid[1]),
                     color=(0, 255, 0), thickness=2)
            cv2.line(frame, (self.pos_centroid[0], self.pos_centroid[1] - 5),
                     (self.pos_centroid[0], self.pos_centroid[1] + 5),
                     color=(0, 255, 0), thickness=2)

            start_index = max(0, len(self.saved_nodes) - 50)
            for i in range(start_index, len(self.saved_nodes)):
                self.annotate_node(frame, point=self.node_pos[i], node=self.saved_nodes[i], t=2)

    def save_to_file(self, fname):
        savelist = []
        # Get the sync time for the frame where the trial ended
        curr_idx = int(self.cap.get(cv2.CAP_PROP_POS_FRAMES)) - 1
        trial_end_sync = self.sync_ts_dict.get(self.ts_column_name, {}).get(curr_idx, self.converted_time)

        with open(fname, 'a+') as file:
            for k, g in groupby(self.saved_nodes):
                savelist.append(k)
            file.writelines('%s,' % items for items in savelist)
            
            file.write('\nSummary Trial {}\n'.format(self.trial_num))
            file.write('Trial End (Sync Seconds): {}\n'.format(trial_end_sync))

            # Debug: end-trial decision + goal-injection outcome
            dbg = getattr(self, '_last_trial_debug', None)
            if dbg is not None:
                file.write(
                    '[DEBUG] reason="{reason}" | goal_node={goal_name} | '
                    'dist_to_goal_px={dist_to_goal_px} (radius={goal_radius_px}) | '
                    'rat_reached_goal={rat_reached_goal} | '
                    'last_node_before={last_node_before_inject} | '
                    'goal_injected={goal_injected}\n'.format(**dbg)
                )

            file.write('Start-Next Nodes // Sync Time (s) // Diff (s) // Length (m) // Velocity (m/s)\n')
            
            for i in range(0, len(self.summary_trial)):
                # summary_trial[i] contains: [(nodes), (times), difference, length, speed]
                line = " ".join(map(str, self.summary_trial[i]))
                file.write(line + '\n')
            file.write('\n')

    def _close_excel_if_open(self, filepath):
        """Quit Excel if it has filepath locked, then wait until the lock releases."""
        import platform
        import subprocess

        def is_locked(path):
            if not os.path.exists(path):
                return False
            try:
                with open(path, 'a'):
                    return False
            except (IOError, PermissionError):
                return True

        if not is_locked(filepath):
            return

        print(f"[POST] '{os.path.basename(filepath)}' is open — closing Excel...")
        system = platform.system()
        if system == 'Darwin':
            subprocess.run(
                ['osascript', '-e', 'tell application "Microsoft Excel" to quit saving no'],
                capture_output=True
            )
        elif system == 'Windows':
            subprocess.run(['taskkill', '/f', '/im', 'EXCEL.EXE'], capture_output=True)

        for _ in range(20):
            time.sleep(0.5)
            if not is_locked(filepath):
                print("[POST] Excel closed successfully.")
                return
        print("[POST] Warning: file may still be locked — proceeding anyway.")

    def post_process_xlsx(self):
        import shutil
        import openpyxl

        if not self.xlsx_src_path or not os.path.exists(self.xlsx_src_path):
            print("[POST] No source xlsx found, skipping post-processing.")
            return

        # --- Copy xlsx to output folder ---
        xlsx_dst = os.path.join(self.out_path, os.path.basename(self.xlsx_src_path))
        self._close_excel_if_open(self.xlsx_src_path)
        shutil.copy2(self.xlsx_src_path, xlsx_dst)
        print(f"[POST] Copied RecordingMeta.xlsx to: {xlsx_dst}")

        # --- Parse txt file for paths per trial ---
        paths_by_trial = {}
        if os.path.exists(self.save):
            with open(self.save, 'r') as f:
                content = f.read()
            current_path = None
            for line in content.splitlines():
                line = line.strip()
                if line.startswith('Summary Trial'):
                    try:
                        trial_num_txt = int(line.split('Summary Trial')[1].strip())
                        if current_path is not None:
                            paths_by_trial[trial_num_txt] = current_path
                        current_path = None
                    except ValueError:
                        pass
                elif (current_path is None and line
                      and not line.startswith('Trial End')
                      and not line.startswith('Start-Next')
                      and not line.startswith('Rat number')):
                    # Node path lines contain only numbers and commas
                    nodes = [n.strip() for n in line.split(',') if n.strip().isdigit()]
                    if nodes:
                        current_path = ','.join(nodes)

        # --- Build lookups by trial_num ---
        delays_by_trial = {tn: d for tn, d in self.trial_delays}
        avg_speed_by_trial = {tn: s for tn, s, _ in self.trial_speed_stats}
        avg_node_speed_by_trial = {tn: ns for tn, _, ns in self.trial_speed_stats}
        start_ts_by_trial = {tn: s for tn, s, _ in self.trial_times}
        end_ts_by_trial = {tn: e for tn, _, e in self.trial_times}

        # --- Write new columns into copied xlsx ---
        try:
            wb = openpyxl.load_workbook(xlsx_dst)
            ws = wb.active

            # Find header row (row 1) and last used column
            headers = [cell.value for cell in ws[1]]
            last_col = len(headers) + 1

            def get_or_add_col(name):
                nonlocal last_col
                if name not in headers:
                    ws.cell(row=1, column=last_col, value=name)
                    col = last_col
                    last_col += 1
                else:
                    col = headers.index(name) + 1
                return col

            path_col          = get_or_add_col('paths')
            delay_col         = get_or_add_col('delay')
            active_time_col   = get_or_add_col('active_time')
            avg_speed_col     = get_or_add_col('avg_speed')
            avg_node_col      = get_or_add_col('avg_between_node_speed')
            start_ts_col      = get_or_add_col('trial_start_time')
            end_ts_col        = get_or_add_col('trial_end_time')

            # Fill rows: row 2 = trial 1, row 3 = trial 2, ...
            num_trials = int(self.num_trials)
            for i in range(num_trials):
                trial_num = i + 1
                row = i + 2
                ws.cell(row=row, column=path_col,        value=paths_by_trial.get(trial_num, ''))
                ws.cell(row=row, column=delay_col,       value=delays_by_trial.get(trial_num, ''))
                ws.cell(row=row, column=active_time_col, value=delays_by_trial.get(trial_num, ''))
                ws.cell(row=row, column=avg_speed_col,   value=avg_speed_by_trial.get(trial_num, ''))
                ws.cell(row=row, column=avg_node_col,    value=avg_node_speed_by_trial.get(trial_num, ''))
                ws.cell(row=row, column=start_ts_col,    value=start_ts_by_trial.get(trial_num, ''))
                ws.cell(row=row, column=end_ts_col,      value=end_ts_by_trial.get(trial_num, ''))

            self._close_excel_if_open(xlsx_dst)
            wb.save(xlsx_dst)
            print(f"[POST] Updated xlsx with paths, active_time, avg_speed, avg_between_node_speed, trial_start/end_time ({num_trials} trials).")
        except Exception as e:
            print(f"[POST] Failed to update xlsx: {e}")

    def find_location(self, start_nodes, goal_nodes):
        nodes_dict = self.nodes_dict
        start_locations = []
        goal_locations = []
        
        for node in start_nodes:
            start_locations.append(nodes_dict.get(str(node)))
            
        for node in goal_nodes:
            goal_locations.append(nodes_dict.get(str(node)))
            
        return start_locations, goal_locations

# --- DATA LOADER ---
def parse_metadata_xlsx(xlsx_path):
    print(f"Reading configuration from: {xlsx_path}")
    try:
        df = pd.read_excel(xlsx_path, engine='openpyxl')
        row0 = df.iloc[0] 
        
        # 1. SCALARS
        start_pt = None
        s_min = float(row0.get('Start_Min', 0))
        s_sec = float(row0.get('Start_Sec', 0))
        if s_min > 0 or s_sec > 0:
            start_pt = (s_min * 60) + s_sec

        # 2. LISTS (Scan columns for per-trial data)
        s_nodes = []
        if 'Start_Nodes' in df.columns:
            s_nodes = df['Start_Nodes'].dropna().astype(int).tolist()
            
        g_nodes = []
        if 'Goal_Node' in df.columns:
            g_nodes = df['Goal_Node'].dropna().astype(int).tolist()
            
        t_types = []
        if 'Trial_Type' in df.columns:
            t_types = df['Trial_Type'].dropna().astype(int).tolist()

        # Special_Trials cells accept either a plain trial number ("3"),
        # or "trial_num@MM:SS" (e.g. "3@5:30") to mark the trial as a
        # scheduled / time-locked trial: its start_node won't trigger until
        # the given session time, and any earlier active trial gets
        # force-ended when that time arrives.
        sp_trials = []
        special_start_seconds = {}  # {trial_num (1-based): seconds_from_session_start}
        if 'Special_Trials' in df.columns:
            for raw in df['Special_Trials'].dropna().tolist():
                s = str(raw).strip()
                if not s:
                    continue
                if '@' in s:
                    try:
                        trial_part, time_part = s.split('@', 1)
                        t_num = int(float(trial_part.strip()))
                        time_part = time_part.strip()
                        if ':' in time_part:
                            mm, ss = time_part.split(':', 1)
                            t_secs = int(mm) * 60 + float(ss)
                        else:
                            t_secs = float(time_part)
                        sp_trials.append(t_num)
                        special_start_seconds[t_num] = t_secs
                    except (ValueError, IndexError) as e:
                        print(f"Warning: bad Special_Trials entry '{s}': {e}")
                else:
                    try:
                        sp_trials.append(int(float(s)))
                    except ValueError:
                        print(f"Warning: bad Special_Trials entry '{s}'")

        did_not_reach = []
        dnr_col = [c for c in df.columns if c.lower() == 'did_not_reach']
        if dnr_col:
            did_not_reach = df[dnr_col[0]].dropna().astype(int).tolist()
        un_dict = {}
        if 'Unnormal_Intervals' in df.columns:
            un_list = df['Unnormal_Intervals'].dropna().astype(str).tolist()
            for item in un_list:
                item = item.strip()
                if ":" in item and "-" in item:
                    parts = item.split(":")
                    try:
                        t_num = int(float(parts[0]))
                        times = parts[1].split("-")
                        un_dict[t_num] = (float(times[0]), float(times[1]))
                    except ValueError:
                        pass

        metadata = {
            'start_point': start_pt,
            'custom_trial': int(float(row0.get('Start_At_Trial_Num', 1))),
            'rat': safe_int_str(row0['Rat_ID']),
            'date': safe_int_str(row0['Date']),
            'repeat': safe_int_str(row0['Repeat']),
            'day': safe_int_str(row0['Day']),
            'session': safe_int_str(row0['Session']),
            'num_trials': safe_int_str(row0['Num_Trials']),
            'start_nodes_list': s_nodes,
            'goal_nodes_list': g_nodes,
            'trial_types_list': t_types,
            'special_trials_list': sp_trials,
            'special_start_seconds': special_start_seconds,
            'did_not_reach_list': did_not_reach,
            'unnormal_intervals': un_dict
        }
        return metadata
    except Exception as e:
        print(f"Error parsing Excel file: {e}")
        raise e

# --- MAIN ---
if __name__ == "__main__":
    try:
        node_list = Path('src/tools/node_list_new.csv').resolve()
        print('\n\nTracker version: v2.11-YOLO11 (Headless / Mass Analysis)\n\n')

        # Argument Parsing
        parser = argparse.ArgumentParser(description="Tracker Headless Mode")
        parser.add_argument('--input_folder', required=True, help="Folder containing 'stitched.mp4' and '*RecordingMeta.xlsx'")
        parser.add_argument('--output_folder', required=True, help="Path to output directory")
        parser.add_argument('--onnx_weight', required=True, help="Path to .pt model file (e.g. yolov11x.pt)")
        
        args = parser.parse_args()
        
        in_p = args.input_folder
        out_p = args.output_folder
        model_path = args.onnx_weight 
        print("Model path:")
        print(model_path)
        
        # 1. Define Video Path
        vid_p = os.path.join(in_p, 'stitched.mp4')
        if not os.path.exists(vid_p):
            print(f"ERROR: Video file not found at: {vid_p}")
            sys.exit(1)

        # 2. Find the meta file
        meta_files = glob.glob(os.path.join(in_p, '*RecordingMeta.xlsx'))
        if not meta_files:
            print(f"ERROR: No file found matching pattern '*RecordingMeta.xlsx' in folder: {in_p}")
            sys.exit(1)
            
        xlsx_file = meta_files[0]
        metadata = parse_metadata_xlsx(xlsx_file)
        metadata['xlsx_src_path'] = xlsx_file
        metadata['input_folder'] = in_p

        # 3. Start Tracker
        tracker = Tracker(vp=vid_p, nl=node_list, out=out_p, metadata=metadata, onnx_weight=model_path)
        
        # Optional renaming
        tracker.change_name_csv(out_p)
        
        # Exit successfully
        print("Done.")
        sys.exit(0)

    except KeyboardInterrupt:
        print("\nProcess interrupted by user.")
        sys.exit(1)
    except Exception as e:
        print(f"\nCRITICAL ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
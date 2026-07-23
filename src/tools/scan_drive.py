"""
Runner step [t]: integrity scan of a raw acquisition drive.

The raw drive is laid out as:

    <root>/
      Rat<N>_<implant>/            e.g. Rat5_491390
        <YYYYMMDD>/                one session per day
          <YYYY-MM-DD_HH-MM-SS>/   12 camera videos eye01..eye12 (.mp4 + .meta)
          <..._pre[.rec]>/         recording folders, one per phase
          <..._maze|awake|hab>/      each holds .rec (+ logger_raw.dat + _merged.rec)
          <..._post[2..]>/

This step checks three things and writes a report to <root>:

  1. VIDEOS play: every *.mp4 has a decodable video stream and a positive
     duration (ffprobe if available, else an mp4 'ftyp'+'moov' atom check).
  2. EPHYS phases: any session that HAS ephys (>=1 .rec file) must have all of
     pre, a task phase (maze / awake / hab), and post.
  3. .rec size: no .rec file (nor logger_raw.dat) may be 0 bytes.

Also flags cheap extra anomalies: cross-rat-named .rec under the wrong Rat
folder, leftover copy temp files (.goutputstream-*, *.tmp), and empty recording
folders.

Outputs, all written to <root>:

    drive_scan.xlsx            every table as a sheet (inventory / issues / files)
                               plus one sheet per animal - open and copy-paste
    drive_scan_report.md       the readable summary
    drive_scan_inventory.md    the readable per-folder file listing

Usage:
    python scan_drive.py --root <drive> [--rat Rat5] [--no-videos] [--deep]
                         [--workers 8] [--ffprobe /path/to/ffprobe]
                         [--no-per-rat]
"""

import os
import re
import sys
import argparse
import subprocess
from pathlib import Path
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

# camera-video subfolder name: 2026-06-24_11-10-42
_CAM_DIR_RE = re.compile(r"^\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2}$")
_DATE_RE = re.compile(r"^\d{8}$")
_RAT_RE = re.compile(r"[Rr]at\s*_?(\d+)")
_TEMP_RE = re.compile(r"(^\.goutputstream-|\.tmp$|\.partial$|\.crdownload$)", re.I)

PHASES = ("pre", "task", "post")

# Folders never worth descending into when hunting for scattered session data.
_SKIP_DIRS = {
    "$recycle.bin", "system volume information", "windows", "winnt",
    "program files", "program files (x86)", "programdata", "appdata",
    "$windows.~bt", "$windows.~ws", "msocache", "recovery", "perflogs",
    "node_modules", "__pycache__", ".git", ".svn", ".trash", ".trashes",
    "library", "applications", "system",
}

# Folder-name fragments that mean "derived output, not raw acquisition" — the
# preprocessing tree (HM_neuron_preprocess) holds DLC coordinates, LFP exports
# and labelled videos for sessions whose raw copies live elsewhere. Scanning it
# only turns up look-alikes of data we already have, so it is skipped whole.
_SKIP_SUBSTRINGS = ("preprocess",)


def skip_dir(name: str) -> bool:
    """True if a directory should never be descended into during a scan — a known
    junk/system folder, a dotfile, or anything in the preprocessing tree."""
    n = name.lower()
    if n in _SKIP_DIRS or name.startswith("."):
        return True
    return any(s in n for s in _SKIP_SUBSTRINGS)


def _classify_phase(name):
    """Map a recording folder / .rec name to pre | task | post | None."""
    n = name.lower()
    if "post" in n:
        return "post"
    if any(k in n for k in ("maze", "mazs", "awake", "hab")):
        return "task"
    if "pre" in n:
        return "pre"
    return None


def _rat_of(name):
    m = _RAT_RE.search(name)
    return int(m.group(1)) if m else None


# A recording name embeds an acquisition timestamp: Rat5_HM_neurons_20260619_160501_post
_REC_TS_RE = re.compile(r"\d{8}_(\d{6})")
# Explicit "resumed after a crash" markers the acquisition software appends when
# a recording is split: ..._post_part2, ..._awake_part2, ..._pre_2, ..._post2.
_SPLIT_SUFFIX_RE = re.compile(r"(_part_?\d+|part_?\d+|_\d+)$", re.I)


def _stem(name):
    """A recording entry's name without a trailing .rec, for phase/split tests."""
    n = name
    for suf in (".rec",):
        if n.lower().endswith(suf):
            n = n[: -len(suf)]
    return n


def detect_session_splits(sess):
    """Which phases of one session were recorded in more than one part.

    When the acquisition PC crashes and the recording is restarted, that phase
    ends up as two files/folders at different clock times — ``..._post`` plus
    ``..._post_part2`` — or simply two recordings the same phase classifies to.
    They are one session's data, not two, and certainly not separate sessions.

    Returns ``{phase: count}`` for phases that have more than one part (so an
    empty dict means "not split"). Looks only at the top-level recording entries,
    which is where each phase's recording lives; reads names only, no contents."""
    # phase -> set of distinct acquisition timestamps seen for it
    times = {}
    forced = set()                       # phases with an explicit _part2 marker
    try:
        entries = list(sess.iterdir())
    except OSError:
        return {}
    for c in entries:
        stem = _stem(c.name)
        ph = _classify_phase(stem)
        if ph is None:
            continue
        # Ignore camera folders (2026-06-22_11-11-46) and merged rollups, which
        # are not themselves a recording of a phase.
        if _CAM_DIR_RE.match(c.name) or c.name.lower().endswith("_merged.rec"):
            continue
        m = _REC_TS_RE.search(c.name)
        ts = m.group(1) if m else stem            # fall back to the whole name
        times.setdefault(ph, set()).add(ts)
        if _SPLIT_SUFFIX_RE.search(stem):
            forced.add(ph)
    out = {}
    for ph, tss in times.items():
        if len(tss) > 1 or ph in forced:
            out[ph] = max(len(tss), 2)            # at least 2 parts if forced
    return out


# ------------------------------------------------------------
#                       video checking
# ------------------------------------------------------------
def _find_ffprobe(explicit):
    """Locate an ffprobe binary: --ffprobe, FFPROBE_CMD, next to FFMPEG_CMD, or
    PATH. Returns the command string, or None if none works."""
    cands = []
    if explicit:
        cands.append(explicit)
    if os.environ.get("FFPROBE_CMD"):
        cands.append(os.environ["FFPROBE_CMD"])
    ff = os.environ.get("FFMPEG_CMD")
    if ff:
        cands.append(re.sub(r"ffmpeg(\.exe)?$", r"ffprobe\1", ff, flags=re.I))
    cands.append("ffprobe")
    for c in cands:
        try:
            subprocess.run([c, "-version"], capture_output=True, timeout=10)
            return c
        except Exception:
            continue
    return None


def _moov_check(path, size):
    """Pure-python fallback: a finalized mp4 has an 'ftyp' box at the start and a
    'moov' box (present iff the file was closed properly, i.e. not truncated)."""
    try:
        with open(path, "rb") as f:
            head = f.read(64)
            if b"ftyp" not in head:
                return False, "no ftyp box (not a finalized mp4?)"
            tail = min(size, 4 * 1024 * 1024)
            f.seek(size - tail)
            if b"moov" in f.read(tail) or b"moov" in head:
                return True, ""
            if size < 128 * 1024 * 1024:      # small file: scan the whole thing
                f.seek(0)
                if b"moov" in f.read():
                    return True, ""
            return False, "no moov atom (truncated / still copying?)"
    except OSError as e:
        return False, f"read failed: {e}"


def check_video(path, ffprobe, deep=False):
    """(ok, reason). ok=True means the file has a decodable video stream."""
    try:
        size = path.stat().st_size
    except OSError as e:
        return False, f"stat failed: {e}"
    if size == 0:
        return False, "0 bytes"
    if not ffprobe:
        return _moov_check(path, size)
    cmd = [ffprobe, "-v", "error", "-select_streams", "v:0",
           "-show_entries", "stream=codec_type:format=duration",
           "-of", "default=noprint_wrappers=1:nokey=1", str(path)]
    try:
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    except subprocess.TimeoutExpired:
        return False, "ffprobe timeout"
    except Exception as e:
        return False, f"ffprobe failed: {e}"
    if r.returncode != 0:
        last = (r.stderr or "").strip().splitlines()
        return False, "ffprobe: " + (last[-1] if last else "error")
    toks = (r.stdout or "").split()
    if "video" not in toks:
        return False, "no video stream"
    dur = next((float(t) for t in toks if _is_float(t)), None)
    if dur is not None and dur <= 0:
        return False, "zero duration"
    if deep:                                  # full decode — catches mid-file corruption
        ff = re.sub(r"ffprobe(\.exe)?$", r"ffmpeg\1", ffprobe, flags=re.I)
        try:
            r2 = subprocess.run([ff, "-v", "error", "-xerror", "-i", str(path),
                                 "-f", "null", "-"], capture_output=True, text=True, timeout=1800)
            if r2.returncode != 0:
                last = (r2.stderr or "").strip().splitlines()
                return False, "decode error: " + (last[-1] if last else "?")
        except subprocess.TimeoutExpired:
            return False, "decode timeout"
    return True, ""


def _is_float(s):
    try:
        float(s)
        return True
    except ValueError:
        return False


# ------------------------------------------------------------
#                       drive scanning
# ------------------------------------------------------------
def find_sessions(root):
    """Session-date folders (<YYYYMMDD>) under `root`, whether root is the drive
    root (root/Rat*/date), a single Rat folder (root/date), or a date folder."""
    if _DATE_RE.match(root.name):
        return [root]
    out = []
    try:
        for child in sorted(root.iterdir()):
            if not child.is_dir():
                continue
            if _DATE_RE.match(child.name):
                out.append(child)                     # root is a Rat folder
            else:
                for g in sorted(child.iterdir()):     # root/Rat*/date
                    if g.is_dir() and _DATE_RE.match(g.name):
                        out.append(g)
    except OSError:
        pass
    return out


def list_drive_roots(include_system=False, include_network=False):
    """Every mounted volume worth searching, as a list of Path.

    Windows: real drive letters, filtered by GetDriveTypeW so CD/DVD and (by
    default) network and system drives are left out. POSIX: the usual removable
    mount points. The system drive is excluded unless asked for — scattered raw
    data lives on the externals, and walking C:\\ is slow for nothing."""
    roots = []
    if sys.platform.startswith("win"):
        import string
        import ctypes
        DRIVE_REMOVABLE, DRIVE_FIXED, DRIVE_REMOTE = 2, 3, 4
        want = {DRIVE_REMOVABLE, DRIVE_FIXED}
        if include_network:
            want.add(DRIVE_REMOTE)
        try:
            k32 = ctypes.windll.kernel32
            mask = k32.GetLogicalDrives()
        except Exception:
            k32, mask = None, None
        sysdrive = os.environ.get("SystemDrive", "C:")[0].upper()
        for i, letter in enumerate(string.ascii_uppercase):
            if mask is not None and not (mask >> i) & 1:
                continue
            p = Path(f"{letter}:\\")
            if mask is None and not p.exists():
                continue
            if k32 is not None and k32.GetDriveTypeW(str(p)) not in want:
                continue
            if not include_system and letter == sysdrive:
                continue
            roots.append(p)
    else:
        for base in ("/Volumes", "/media", "/mnt", "/run/media"):
            b = Path(base)
            if not b.is_dir():
                continue
            try:
                for child in sorted(b.iterdir()):
                    if child.is_dir():
                        # /media/<user>/<label> on many Linux setups
                        subs = [g for g in child.iterdir() if g.is_dir()] \
                            if base in ("/media", "/run/media") else []
                        roots += subs or [child]
            except OSError:
                continue
    return roots


def find_sessions_deep(root, max_depth=6, on_dir=None, should_stop=None):
    """Find every ``<Rat folder>/<YYYYMMDD>`` session under `root`, at any depth
    up to `max_depth` — unlike find_sessions(), which only looks one or two
    levels down from a known drive root.

    Use this to locate data that was filed somewhere unexpected (a backup
    subfolder, a per-experimenter folder, a nested copy). `on_dir` is called with
    each directory visited (progress); `should_stop` is polled to allow cancel."""
    root = Path(root)
    out = []
    stack = [(root, 0)]
    while stack:
        if should_stop is not None and should_stop():
            break
        d, depth = stack.pop()
        if on_dir is not None:
            on_dir(d)
        try:
            children = [c for c in d.iterdir() if c.is_dir()]
        except OSError:
            continue                       # unreadable / permission denied
        for c in children:
            if skip_dir(c.name):
                continue
            if _DATE_RE.match(c.name):
                # a session folder — only counts if its parent names a rat, and
                # never worth descending into either way
                if _rat_of(d.name) is not None:
                    out.append(c)
                continue
            if depth < max_depth:
                stack.append((c, depth + 1))
    return out


def _classify_file(name):
    """File kind for the inventory: video / meta / merged / rec / logger / config / other."""
    n = name.lower()
    if n.endswith(".mp4"):
        return "video"
    if n.endswith(".meta"):
        return "meta"
    if n.endswith("_merged.rec"):
        return "merged"
    if n.endswith(".rec"):
        return "rec"
    if n == "logger_raw.dat":
        return "logger"
    if n.endswith(".trodesconf"):
        return "config"
    return "other"


def scan_session(sess, issues, inv_rows, file_rows):
    """Check one session (ephys-phase completeness + zero-size .rec + cross-rat +
    junk) AND inventory it. Appends issues, one inventory row (per session), and
    one file row per file. Returns (has_ephys, mp4_list)."""
    rat_dir = sess.parent.name
    rat_no = _rat_of(rat_dir)
    mp4s, rec_files, phases, rec_folders, empty_dirs = [], [], set(), [], []
    n = {k: 0 for k in ("video", "meta", "rec", "merged", "logger", "config", "other")}
    bytes_video = bytes_ephys = 0
    phase_bytes = {p: 0 for p in PHASES}   # ephys bytes per pre/task/post

    # Groups of files to inventory: (folder_label, files, is_recording). Some rats
    # (Rat3/Rat4) drop the camera videos LOOSE in the date folder (no timestamp
    # subfolder), so scan those direct files too — else they'd be missed entirely.
    groups = []
    loose = [p for p in sess.iterdir() if p.is_file()]
    if loose:
        groups.append(("(date folder)", loose, False))
    for sub in sorted(p for p in sess.iterdir() if p.is_dir()):
        entries = list(sub.iterdir())
        if not entries:
            empty_dirs.append(sub)
        is_cam = _CAM_DIR_RE.match(sub.name) or any(p.suffix.lower() == ".mp4" for p in entries)
        groups.append((sub.name, [p for p in sub.rglob("*") if p.is_file()], not is_cam))

    for folder, files, is_recording in groups:
        has_rec_here = False
        here_ephys = 0                       # ephys bytes in this recording folder
        for p in files:
            cat = _classify_file(p.name)
            try:
                size = p.stat().st_size
            except OSError:
                size = -1
            n[cat] += 1
            file_rows.append(dict(rat=rat_dir, session=sess.name, folder=folder,
                                  file=p.name, type=cat, size_bytes=size))
            if cat == "video":
                mp4s.append(p); bytes_video += max(size, 0)
            elif cat in ("rec", "merged", "logger"):
                rec_files.append(p); bytes_ephys += max(size, 0); has_rec_here = True
                here_ephys += max(size, 0)
                if cat in ("rec", "merged"):
                    rno = _rat_of(p.name)
                    if rat_no is not None and rno is not None and rno != rat_no:
                        issues.append(dict(category="cross-rat", rat=rat_dir, session=sess.name,
                                           path=str(p), detail=f"Rat{rno} file under {rat_dir}"))
            if _TEMP_RE.search(p.name):
                issues.append(dict(category="leftover-temp", rat=rat_dir, session=sess.name,
                                   path=str(p), detail="copy temp / partial file"))
        if is_recording and has_rec_here:
            ph = _classify_phase(folder)
            if ph:
                phases.add(ph)
                phase_bytes[ph] += here_ephys
            rec_folders.append(f"{folder}[{ph or '?'}]")

    has_ephys = bool(rec_files)

    # (3) zero-byte .rec / logger
    for p in rec_files:
        try:
            if p.stat().st_size == 0:
                issues.append(dict(category="zero-size-rec", rat=rat_dir, session=sess.name,
                                   path=str(p), detail="0 bytes"))
        except OSError as e:
            issues.append(dict(category="stat-failed", rat=rat_dir, session=sess.name,
                               path=str(p), detail=str(e)))
    for d in empty_dirs:
        issues.append(dict(category="empty-folder", rat=rat_dir, session=sess.name,
                           path=str(d), detail="folder has no files"))
    # (2) ephys phase completeness — only enforced when the session has ephys
    if has_ephys:
        missing = [p for p in PHASES if p not in phases]
        if missing:
            issues.append(dict(category="missing-phase", rat=rat_dir, session=sess.name,
                               path=str(sess),
                               detail=f"missing {'+'.join(missing)} (have {'+'.join(sorted(phases)) or 'none'})"))

    inv_rows.append(dict(
        rat=rat_dir, session=sess.name, has_ephys=int(has_ephys),
        phases="+".join(p for p in PHASES if p in phases) if has_ephys else "-",
        n_video=n["video"], n_meta=n["meta"],
        n_rec=n["rec"], n_merged=n["merged"], n_logger=n["logger"],
        video_gb=round(bytes_video / 1e9, 2), ephys_gb=round(bytes_ephys / 1e9, 2),
        phase_bytes=dict(phase_bytes),
        rec_folders="; ".join(rec_folders)))
    return has_ephys, mp4s


def run(root, do_videos=True, deep=False, workers=8, rat_filter=None, ffprobe_cmd=None,
        paste_per_rat=False):
    root = Path(root)
    sessions = find_sessions(root)
    if rat_filter:
        sessions = [s for s in sessions if rat_filter.lower() in s.parent.name.lower()]
    if not sessions:
        print(f"No <YYYYMMDD> session folders found under {root}.")
        return
    print(f"Found {len(sessions)} session(s) under {root}.")

    issues, inv_rows, file_rows = [], [], []
    all_mp4s = []
    n_ephys = 0
    for sess in sessions:
        try:
            has_ephys, mp4s = scan_session(sess, issues, inv_rows, file_rows)
        except OSError as e:
            issues.append(dict(category="scan-failed", rat=sess.parent.name,
                               session=sess.name, path=str(sess), detail=str(e)))
            continue
        n_ephys += int(has_ephys)
        all_mp4s += mp4s

    # (1) video playability — threaded (ffprobe is subprocess/I-O bound)
    bad_videos = []
    if do_videos and all_mp4s:
        ffprobe = _find_ffprobe(ffprobe_cmd)
        mode = (f"ffprobe{' + full decode' if deep else ''}"
                if ffprobe else "python moov-atom check (ffprobe not found)")
        print(f"Checking {len(all_mp4s)} video(s) [{mode}] ...")
        with ThreadPoolExecutor(max_workers=max(1, workers)) as ex:
            futs = {ex.submit(check_video, p, ffprobe, deep): p for p in all_mp4s}
            done = 0
            for fut in as_completed(futs):
                p = futs[fut]
                done += 1
                if done % 200 == 0:
                    print(f"  ... {done}/{len(all_mp4s)}")
                ok, reason = fut.result()
                if not ok:
                    rat = p.parent.parent.parent.name if len(p.parts) >= 3 else ""
                    bad_videos.append(dict(category="bad-video", rat=rat, session="",
                                           path=str(p), detail=reason))
        issues += bad_videos

    _write_inventory(root, inv_rows, file_rows, per_rat=paste_per_rat)
    _write_workbook(root, inv_rows, file_rows, issues, per_rat=paste_per_rat)
    _write_report(root, sessions, n_ephys, len(all_mp4s), bad_videos, issues, do_videos, inv_rows)


def _hsize(n):
    """Human-readable byte size."""
    if n is None or n < 0:
        return "?"
    for unit in ("B", "KB", "MB", "GB", "TB"):
        if n < 1024 or unit == "TB":
            return f"{n:.0f} {unit}" if unit == "B" else f"{n:.1f} {unit}"
        n /= 1024.0


def _write_inventory_md(root, inv_rows, file_rows):
    """Readable per-animal / per-session / per-folder file listing. Camera folders
    are collapsed to one line (12x eye*.mp4); recording folders list each file
    with its size. Zero-byte and cross-rat files are flagged inline."""
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # index file rows by (rat, session, folder), preserving insertion order
    per_folder = {}
    for r in file_rows:
        per_folder.setdefault((r["rat"], r["session"], r["folder"]), []).append(r)
    inv_by = {(r["rat"], r["session"]): r for r in inv_rows}
    rats = sorted({r["rat"] for r in inv_rows})

    lines = [f"# Drive inventory — {root}", f"_{ts}_", ""]
    for rat in rats:
        sessions = sorted({s for (rr, s) in inv_by if rr == rat})
        ne = sum(inv_by[(rat, s)]["has_ephys"] for s in sessions)
        lines.append(f"## {rat} — {len(sessions)} session(s), {ne} with ephys")
        for s in sessions:
            iv = inv_by[(rat, s)]
            head = iv["phases"] if iv["has_ephys"] else "no ephys"
            lines.append(f"### {s} — {head}  ·  ephys {iv['ephys_gb']} GB  ·  video {iv['video_gb']} GB")
            folders = [k[2] for k in per_folder if k[0] == rat and k[1] == s]
            for folder in folders:
                rows = per_folder[(rat, s, folder)]
                vids = [x for x in rows if x["type"] == "video"]
                metas = [x for x in rows if x["type"] == "meta"]
                rest = [x for x in rows if x["type"] not in ("video", "meta")]
                if vids:                                   # camera folder — collapse
                    tot = sum(max(x["size_bytes"], 0) for x in vids)
                    bad0 = sum(1 for x in vids if x["size_bytes"] == 0)
                    extra = f", {len(metas)}x .meta" if metas else ""
                    warn = f"  ⚠ {bad0} zero-byte" if bad0 else ""
                    lines.append(f"- 📹 `{folder}` — {len(vids)}x video ({_hsize(tot)}){extra}{warn}")
                    for x in rest:
                        lines.append(f"    - `{x['file']}` — {_hsize(x['size_bytes'])}")
                else:                                      # recording folder — list files
                    ph = _classify_phase(folder)
                    lines.append(f"- 🧠 `{folder}`" + (f"  _[{ph}]_" if ph else "  _[phase?]_"))
                    for x in sorted(rows, key=lambda z: z["file"]):
                        flags = ""
                        if x["size_bytes"] == 0:
                            flags += "  ⚠ 0 bytes"
                        rno = _rat_of(x["file"])
                        rrno = _rat_of(rat)
                        if x["type"] in ("rec", "merged") and rrno and rno and rno != rrno:
                            flags += f"  ⚠ Rat{rno} file"
                        lines.append(f"    - `{x['file']}` — {_hsize(x['size_bytes'])}{flags}")
            lines.append("")
    md_path = root / "drive_scan_inventory.md"
    try:
        md_path.write_text("\n".join(lines), encoding="utf-8")
    except OSError as e:
        print(f"Could not write {md_path} ({e}).")
    return md_path


INV_COLS = ["rat", "session", "has_ephys", "phases", "n_video", "n_meta",
            "n_rec", "n_merged", "n_logger", "video_gb", "ephys_gb", "rec_folders"]


ISSUE_COLS = ["category", "rat", "session", "path", "detail"]
FILE_COLS = ["rat", "session", "folder", "file", "type", "size_bytes"]

# Excel forbids these in a sheet name, and caps the name at 31 characters.
_BAD_SHEET = re.compile(r"[\[\]:*?/\\]")


def _sheet_name(name, taken):
    """A legal, unique Excel sheet name for `name`."""
    s = _BAD_SHEET.sub("_", str(name)).strip("'") or "sheet"
    s = s[:31]
    base, i = s, 2
    while s.lower() in taken:
        suffix = f"_{i}"
        s = base[:31 - len(suffix)] + suffix
        i += 1
    taken.add(s.lower())
    return s


def _add_sheet(wb, title, cols, rows, taken):
    """One table as a worksheet: bold frozen header, then the rows, columns sized
    to their content so nothing has to be widened before reading or copying."""
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    ws = wb.create_sheet(_sheet_name(title, taken))
    ws.append(list(cols))
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.freeze_panes = "A2"
    widths = [len(str(c)) for c in cols]
    for r in rows:
        vals = [r.get(c) for c in cols]
        # dicts and lists have no cell representation; the phase_bytes column is
        # the one that would otherwise raise here
        vals = [v if v is None or isinstance(v, (str, int, float)) else str(v)
                for v in vals]
        ws.append(vals)
        for i, v in enumerate(vals):
            widths[i] = max(widths[i], min(len(str(v)) if v is not None else 0, 60))
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w + 2
    return ws


def write_xlsx(path, sheets):
    """Write tables to an .xlsx. `sheets` is [(title, cols, rows-as-dicts), ...].

    Shared with scan_drive_gui so its export buttons produce the same workbook
    styling as the scan itself. Returns (ok, message).
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        return False, "openpyxl is not installed (pip install openpyxl)."
    wb = Workbook()
    wb.remove(wb.active)
    taken = set()
    for title, cols, rows in sheets:
        _add_sheet(wb, title, cols, rows, taken)
    try:
        wb.save(path)
    except OSError as e:
        return False, str(e)
    return True, str(path)


def _write_workbook(root, inv_rows, file_rows, issues, per_rat=False):
    """Every table of the scan in one .xlsx, so the results can be opened and
    copied straight into a spreadsheet without an import step. The CSV/TSV files
    stay as they are for anything that reads them programmatically."""
    try:
        from openpyxl import Workbook
    except ImportError:
        print("openpyxl not installed - skipping drive_scan.xlsx (pip install openpyxl).")
        return
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    xlsx = root / "drive_scan.xlsx"
    wb = Workbook()
    wb.remove(wb.active)                       # drop the default empty sheet
    taken = set()
    inv_sorted = sorted(inv_rows, key=lambda x: (x["rat"], x["session"]))
    _add_sheet(wb, "inventory", INV_COLS + ["scanned_at"],
               [{**r, "scanned_at": ts} for r in inv_sorted], taken)
    _add_sheet(wb, "issues", ISSUE_COLS, issues, taken)
    _add_sheet(wb, "files", FILE_COLS, file_rows, taken)
    if per_rat:
        by_rat = {}
        for r in inv_sorted:
            by_rat.setdefault(r["rat"], []).append(r)
        for rat, rows in sorted(by_rat.items()):
            _add_sheet(wb, rat, INV_COLS + ["scanned_at"],
                       [{**r, "scanned_at": ts} for r in rows], taken)
    try:
        wb.save(xlsx)
    except OSError as e:
        print(f"Could not write {xlsx} ({e}).")
        return
    print(f"Excel workbook: {xlsx}")
    print(f"  sheets: {', '.join(wb.sheetnames)}")


def _write_inventory(root, inv_rows, file_rows, per_rat=False):
    """The readable per-folder file listing. Every tabular output now lives in
    drive_scan.xlsx, written separately by _write_workbook."""
    _write_inventory_md(root, inv_rows, file_rows)


def _write_report(root, sessions, n_ephys, n_videos, bad_videos, issues, did_videos, inv_rows=None):
    order = ["missing-phase", "zero-size-rec", "bad-video", "cross-rat",
             "empty-folder", "leftover-temp", "stat-failed", "scan-failed"]
    by_cat = {c: [i for i in issues if i["category"] == c] for c in order}
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Markdown report (human-readable)
    titles = {
        "missing-phase": "Sessions with ephys MISSING a phase (pre / task / post)",
        "zero-size-rec": "Zero-byte .rec / logger files",
        "bad-video": "Unplayable / truncated videos",
        "cross-rat": "Cross-rat-named files (wrong Rat folder)",
        "empty-folder": "Empty recording folders",
        "leftover-temp": "Leftover copy temp files (.goutputstream / .tmp)",
        "stat-failed": "Files that could not be stat'd",
        "scan-failed": "Sessions that could not be scanned",
    }
    inv_rows = inv_rows or []
    rats = sorted({r["rat"] for r in inv_rows})
    lines = [f"# Drive scan — {root}", f"_{ts}_", "",
             f"- Animals: **{len(rats)}** ({', '.join(rats) if rats else '-'})",
             f"- Sessions: **{len(sessions)}** (with ephys: **{n_ephys}**)",
             f"- Videos checked: **{n_videos if did_videos else 'skipped'}**"
             + (f" (bad: **{len(bad_videos)}**)" if did_videos else ""),
             f"- Total issues: **{len(issues)}**", ""]

    # Inventory: which animals, which sessions, and what's in each
    lines.append("## Inventory — animals / sessions / files")
    by_rat = {}
    for r in inv_rows:
        by_rat.setdefault(r["rat"], []).append(r)
    for rat in rats:
        rows = sorted(by_rat[rat], key=lambda x: x["session"])
        ne = sum(x["has_ephys"] for x in rows)
        lines.append(f"### {rat} — {len(rows)} session(s), {ne} with ephys")
        lines.append("| session | phases | videos | rec/merged/logger | ephys GB | video GB |")
        lines.append("|---|---|---|---|---|---|")
        for x in rows:
            lines.append(f"| {x['session']} | {x['phases']} | {x['n_video']} "
                         f"| {x['n_rec']}/{x['n_merged']}/{x['n_logger']} "
                         f"| {x['ephys_gb']} | {x['video_gb']} |")
        lines.append("")
    lines.append("_Readable per-file listing: **drive_scan_inventory.md** · "
                 "every table as a sheet: **drive_scan.xlsx**_")
    lines.append("")

    for c in order:
        rows = by_cat[c]
        if not rows and c in ("stat-failed", "scan-failed"):
            continue
        lines.append(f"## {titles[c]} — {len(rows)}")
        for i in rows[:500]:
            loc = f"{i['rat']}/{i['session']}" if i["session"] else i["rat"]
            lines.append(f"- `{loc}` — {i['detail']}  \n  `{i['path']}`")
        if len(rows) > 500:
            lines.append(f"- … and {len(rows) - 500} more (see CSV)")
        lines.append("")
    md_path = root / "drive_scan_report.md"
    try:
        md_path.write_text("\n".join(lines), encoding="utf-8")
    except OSError as e:
        print(f"Could not write {md_path} ({e}).")

    # Console summary
    print("\n" + "=" * 56)
    print(f"DRIVE SCAN: {len(rats)} animals, {len(sessions)} sessions "
          f"({n_ephys} with ephys), {len(issues)} issue(s)")
    for c in order:
        if by_cat[c]:
            print(f"  {c:15s}: {len(by_cat[c])}")
    print(f"Report:    {md_path}")
    print(f"Inventory: {root / 'drive_scan_inventory.md'}  (readable file listing)")
    print(f"Excel:     {root / 'drive_scan.xlsx'}  (inventory / issues / files + one sheet per animal)")
    print("=" * 56)


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Integrity scan of a raw acquisition drive (step t).")
    ap.add_argument("--root", required=True, help="drive / folder to scan (drive root, a Rat folder, or a date folder).")
    ap.add_argument("--config", default=None, help="accepted for runner consistency (unused).")
    ap.add_argument("--rat", default=None, help="only scan sessions under Rat folders matching this substring.")
    ap.add_argument("--no-videos", dest="videos", action="store_false", help="skip the video-playability check.")
    ap.add_argument("--deep", action="store_true", help="fully decode each video (slow) instead of a header check.")
    ap.add_argument("--workers", type=int, default=8, help="parallel video checks (default 8).")
    ap.add_argument("--ffprobe", default=None, help="path to ffprobe (else FFPROBE_CMD / FFMPEG_CMD sibling / PATH).")
    ap.add_argument("--no-per-rat", dest="paste_per_rat", action="store_false",
                    help="skip the per-animal sheets in the workbook.")
    args = ap.parse_args()
    try:
        run(args.root, do_videos=args.videos, deep=args.deep, workers=args.workers,
            rat_filter=args.rat, ffprobe_cmd=args.ffprobe,
            paste_per_rat=args.paste_per_rat)
    except Exception as e:
        import traceback
        print(f"[scan-drive] Failed: {e}")
        traceback.print_exc()
        sys.exit(1)
